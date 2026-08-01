"""
Excel-Driven Interactive Fillable PDF Form Generator Engine
============================================================
This module is responsible for reading corporate form specifications from the master Excel
Form Configurator workbook ('Interactive_PDF_Code/fillable_pdf_configurator.xlsx') and
programmatically compiling interactive, professional Fillable PDF forms saved into the
'Interactive_PDF_Code/Interactive_PDF/' subfolder.

Output PDF Filename Standard:
----------------------------
The PDF is automatically named using the 'Form_Title' set in Sheet 1 of the Excel Configurator
suffixed with the current datetime timestamp:
'Interactive_PDF_Code/Interactive_PDF/{Form_Title}_{YYYYMMDD_HHMMSS}.pdf'

Directory Architecture:
-----------------------
- Root Folder: Interactive_PDF_Code/
- Master Excel Configurator: Interactive_PDF_Code/fillable_pdf_configurator.xlsx
- Builder Script: Interactive_PDF_Code/build_pdf_configurator_excel.py
- Generator Engine Script: Interactive_PDF_Code/pdf_generator_from_excel.py
- Generated PDFs Subfolder: Interactive_PDF_Code/Interactive_PDF/

LESSONS LEARNED & ARCHITECTURAL BEST PRACTICES FOR FUTURE DEVELOPMENT:
----------------------------------------------------------------------
1. AcroForm Button Appearance Stream (/AP) Preservation:
   - NEVER delete or remove the '/AP' dictionary entry from an AcroForm button object.
   - Deleting '/AP' causes the button background fill, borders, and hand cursor to become completely invisible in Adobe Acrobat Reader DC.
   - To achieve clean text centering inside buttons, use `_center_pad_button_label()` to calculate mathematical leading space padding. This visually centers text while keeping ReportLab's `/AP` vector stream (background fills `#FFF2CC`/`#1F4E78`, borders, and pointer hand cursors) 100% visible and functional.

2. Acrobat JavaScript RichText Multi-Color Formatting:
   - To render multi-color text (such as bright red validation warnings alongside black summary text) inside an AcroForm text field in Adobe Reader DC:
     a) Inject `/DS` (Default Style) and `/RV` (RichValue XML) streams into the field dictionary object upon creation (`fieldFlags='multiline richText'`).
     b) In Acrobat JavaScript, pass native Acrobat `color.red` and `color.black` Span objects to both `targetField.richValue` and `targetField.richContents`.
     c) Do NOT assign plain text `targetField.value = ...` after `richValue`, as plain text assignment overrides and wipes out rich text formatting in Adobe Reader DC.

3. Dynamic Form Lock/Unlock Synchronization:
   - Initial button caption (`[ UNLOCKED - LOCK EDITING ]` vs `[ LOCKED - UNLOCK EDITING ]`) and fill color (`#1F4E78` Navy vs `#C00000` Red) MUST evaluate `config.lock_status` dynamically on PDF creation so button state matches initial field readability.

Author: Interactive PDF Form Compiler Pipeline
Standards Spec: standards/excel_form_configurator_contracts.md
"""

import io
import os
import sys
import logging
import datetime
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import HexColor
from reportlab.pdfbase import pdfdoc

logger = logging.getLogger(__name__)


@dataclass
class ExcelFormConfig:
    """
    Data model storing global form configuration settings parsed from Sheet 1 ('Form_Basic_Details').
    """
    title: str = "Employee_Information_Form"
    subtitle: str = "OFFICIAL EMPLOYEE DETAILS & RECORD FORM"
    instruction_1: str = "!! Save As - this PDF file before mailing. Cross-check whether all filled form data is properly saved before sending. !!"
    instruction_2: str = "!! Before saving the file Click 'Generate Summary' Button to update it..!!   !! Use Adobe Acrobat Reader software for filling this form. !!"
    instruction_3: str = "3. Verify all employee personal details & department info before generating summary."
    instruction_4: str = ""
    footer_text: str = "Official Employee Information & Department Record Form"
    lock_status: str = "Unlocked"
    navy_primary: str = "#1F4E78"
    soft_blue: str = "#DDEBF7"
    soft_yellow: str = "#FFF2CC"
    summary_section_title: str = "GENERATE FORM SUMMARY & COPY TEXT"
    summary_button_text: str = ">>>> Click the Button - To Generate Form Summary <<<<"
    summary_field_label: str = "GENERATED FORM SUMMARY TEXT"
    page_width: float = 612.0
    page_height: float = 792.0


@dataclass
class ExcelFieldSpec:
    """
    Data model storing individual field specifications parsed from Sheet 3 ('Field_Configuration').
    """
    field_id: str
    section_id: str
    field_label: str
    field_datatype: str
    default_value: str
    tooltip: str
    validation_operator: str
    validation_param: str
    dropdown_options: List[str] = field(default_factory=list)


@dataclass
class ExcelSectionSpec:
    """
    Data model storing individual section layout specifications parsed from Sheet 2 ('Section_Details').
    """
    section_num: int
    section_id: str
    section_title: str
    grid_cols: int = 3


class ExcelConfigReader:
    """
    Parser engine responsible for loading, validating, and extracting form configurations
    from the master Excel Form Configurator workbook.
    """

    @staticmethod
    def read_config(excel_path: str) -> Tuple[ExcelFormConfig, List[ExcelSectionSpec], List[ExcelFieldSpec]]:
        abs_path = os.path.abspath(excel_path)
        if not os.path.exists(abs_path):
            raise FileNotFoundError(f"Master Excel Form Configurator workbook not found at: '{abs_path}'")

        wb = openpyxl.load_workbook(abs_path, data_only=True)
        config = ExcelFormConfig()

        # 1. Parse Sheet 1: Form_Basic_Details
        cfg_sheet_name = "Form_Basic_Details" if "Form_Basic_Details" in wb.sheetnames else ("Form_Config" if "Form_Config" in wb.sheetnames else None)
        if cfg_sheet_name:
            ws_cfg = wb[cfg_sheet_name]
            for row in ws_cfg.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                key = str(row[0]).strip()
                val = str(row[1] or '').strip()

                if key == "Form_Title": config.title = val
                elif key == "Form_Subtitle": config.subtitle = val
                elif key == "Instruction_Point_1": config.instruction_1 = val
                elif key == "Instruction_Point_2": config.instruction_2 = val
                elif key == "Instruction_Point_3": config.instruction_3 = val
                elif key == "Instruction_Point_4": config.instruction_4 = val
                elif key == "Footer_Text": config.footer_text = val
                elif key == "Default_Lock_Status" and val: config.lock_status = val
                elif key == "Theme_Navy_Primary" and val: config.navy_primary = val
                elif key == "Theme_Soft_Blue" and val: config.soft_blue = val
                elif key == "Theme_Soft_Yellow" and val: config.soft_yellow = val
                elif key == "Summary_Section_Title" and val: config.summary_section_title = val
                elif key == "Summary_Button_Text" and val: config.summary_button_text = val
                elif key == "Summary_Field_Label" and val: config.summary_field_label = val

        # 2. Parse Sheet 2: Section_Details
        sec_sheet_name = "Section_Details" if "Section_Details" in wb.sheetnames else None
        section_specs: List[ExcelSectionSpec] = []
        if sec_sheet_name:
            ws_sec = wb[sec_sheet_name]
            for row in ws_sec.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                try:
                    s_num = int(row[0])
                except Exception:
                    s_num = len(section_specs) + 1
                s_id = str(row[1] or '').strip()
                s_title = str(row[2] or '').strip()
                # Ignore legacy summary section rows if present in Excel
                if s_id.upper() == "SEC_5" or "GENERATE SHORT DPR" in s_title.upper():
                    continue
                try:
                    s_cols = int(row[3]) if row[3] else 3
                except Exception:
                    s_cols = 3
                section_specs.append(ExcelSectionSpec(s_num, s_id, s_title, s_cols))

        # 3. Parse Sheet 3: Field_Configuration (Clean 9-Column Architecture)
        sch_sheet_name = "Field_Configuration" if "Field_Configuration" in wb.sheetnames else ("Field_Schema" if "Field_Schema" in wb.sheetnames else None)
        field_specs: List[ExcelFieldSpec] = []

        if sch_sheet_name:
            ws_sch = wb[sch_sheet_name]
            for row in ws_sch.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                f_id = str(row[0]).strip()
                sec_id = str(row[1] or '').strip()
                label = str(row[2] or '').strip()
                # Ignore legacy summary field rows if present in Excel
                if f_id.upper() == "FIELD_5_1" or sec_id.upper() == "SEC_5" or "GENERATED SHORT DPR" in label.upper():
                    continue
                f_dtype = str(row[3] or 'short_text').strip().lower()
                def_val = str(row[4] or '').strip()
                tooltip = str(row[5] or '').strip()
                val_op = str(row[6] or 'none').strip().lower()
                val_param = str(row[7] or '').strip()
                opts_str = str(row[8] or '').strip() if len(row) > 8 else ''



                if f_id.upper() == "FIELD_4_1" and "ENTER PROGRESS SUMMARY" in label.upper():
                    label = "OPERATIONAL PROGRESS NARRATIVE"

                opts = [o.strip() for o in opts_str.split(',') if o.strip()] if opts_str else []

                spec = ExcelFieldSpec(
                    field_id=f_id,
                    section_id=sec_id,
                    field_label=label,
                    field_datatype=f_dtype,
                    default_value=def_val,
                    tooltip=tooltip,
                    validation_operator=val_op,
                    validation_param=val_param,
                    dropdown_options=opts
                )
                field_specs.append(spec)

        logger.info(f"Successfully read Form Configuration ({len(section_specs)} sections, {len(field_specs)} fields) from Master Excel: '{abs_path}'")
        return config, section_specs, field_specs


class AcrobatActionBuilder:
    """
    Constructs native Adobe Acrobat JavaScript action dictionaries (pdfdoc.PDFDictionary)
    for live PDF field validations, auto-calculations, lock toggling, and summary generation.
    """

    @staticmethod
    def _wrap_aa_events(js_code: str) -> pdfdoc.PDFDictionary:
        """Wraps JS code in PDF dictionary listening to F (Format), V (Validate), and Bl (OnBlur)."""
        js_act = pdfdoc.PDFDictionary({
            'S': pdfdoc.PDFName('JavaScript'),
            'JS': pdfdoc.PDFString(js_code)
        })
        return pdfdoc.PDFDictionary({
            'F': js_act,
            'V': js_act,
            'Bl': js_act
        })

    @staticmethod
    def build_whole_number_validation_action() -> pdfdoc.PDFDictionary:
        """Returns Acrobat JavaScript action for live Whole Number (Integer) validation with red highlight."""
        js_code = (
            "var val = event.value;\n"
            "var isInvalid = false;\n"
            "if (val !== undefined && val !== null && (val + '').trim() !== '') {\n"
            "    var n = Number(val);\n"
            "    if (isNaN(n) || !Number.isInteger(n)) {\n"
            "        isInvalid = true;\n"
            "    }\n"
            "}\n"
            "if (isInvalid) {\n"
            "    event.target.fillColor = ['RGB', 1.0, 0.78, 0.81];\n"
            "    event.target.textColor = ['RGB', 0.61, 0.0, 0.02];\n"
            "} else {\n"
            "    event.target.fillColor = ['RGB', 0.98, 0.98, 0.98];\n"
            "    event.target.textColor = ['RGB', 0.0, 0.0, 0.0];\n"
            "}\n"
        )
        return AcrobatActionBuilder._wrap_aa_events(js_code)

    @staticmethod
    def build_decimal_validation_action() -> pdfdoc.PDFDictionary:
        """Returns Acrobat JavaScript action for live Decimal numeric validation with red highlight."""
        js_code = (
            "var val = event.value;\n"
            "var isInvalid = false;\n"
            "if (val !== undefined && val !== null && (val + '').trim() !== '') {\n"
            "    var n = Number(val);\n"
            "    if (isNaN(n)) {\n"
            "        isInvalid = true;\n"
            "    }\n"
            "}\n"
            "if (isInvalid) {\n"
            "    event.target.fillColor = ['RGB', 1.0, 0.78, 0.81];\n"
            "    event.target.textColor = ['RGB', 0.61, 0.0, 0.02];\n"
            "} else {\n"
            "    event.target.fillColor = ['RGB', 0.98, 0.98, 0.98];\n"
            "    event.target.textColor = ['RGB', 0.0, 0.0, 0.0];\n"
            "}\n"
        )
        return AcrobatActionBuilder._wrap_aa_events(js_code)

    @staticmethod
    def build_date_validation_action() -> pdfdoc.PDFDictionary:
        """Returns Acrobat JavaScript action for live visual DD-MM-YYYY date format validation with red highlight."""
        js_code = (
            "var val = event.value;\n"
            "var isInvalid = false;\n"
            "if (val !== undefined && val !== null && (val + '').trim() !== '') {\n"
            "    var parts = (val + '').trim().split('-');\n"
            "    if (parts.length !== 3) { isInvalid = true; }\n"
            "    else {\n"
            "        var d = parseInt(parts[0], 10);\n"
            "        var m = parseInt(parts[1], 10);\n"
            "        var y = parseInt(parts[2], 10);\n"
            "        if (isNaN(d) || isNaN(m) || isNaN(y) || m < 1 || m > 12 || d < 1 || d > 31 || y < 1900 || y > 2100) {\n"
            "            isInvalid = true;\n"
            "        }\n"
            "    }\n"
            "}\n"
            "if (isInvalid) {\n"
            "    event.target.fillColor = ['RGB', 1.0, 0.78, 0.81];\n"
            "    event.target.textColor = ['RGB', 0.61, 0.0, 0.02];\n"
            "} else {\n"
            "    event.target.fillColor = ['RGB', 0.98, 0.98, 0.98];\n"
            "    event.target.textColor = ['RGB', 0.0, 0.0, 0.0];\n"
            "}\n"
        )
        return AcrobatActionBuilder._wrap_aa_events(js_code)

    @staticmethod
    def build_closing_hsd_calc_action() -> pdfdoc.PDFDictionary:
        """Returns Acrobat JavaScript calculate action for Net Closing HSD Stock."""
        js_code = (
            "var st = parseFloat(this.getField('Stock_HSD_KL').value);\n"
            "var cs = parseFloat(this.getField('Consumption_HSD_KL').value);\n"
            "if (!isNaN(st) && !isNaN(cs)) {\n"
            "    event.value = (st - cs).toFixed(1);\n"
            "} else {\n"
            "    event.value = '';\n"
            "}\n"
        )
        return pdfdoc.PDFDictionary({
            'C': pdfdoc.PDFDictionary({
                'S': pdfdoc.PDFName('JavaScript'),
                'JS': pdfdoc.PDFString(js_code)
            })
        })

    @staticmethod
    def build_lock_toggle_action(field_specs: List['ExcelFieldSpec']) -> pdfdoc.PDFDictionary:
        """Returns dynamic Acrobat JavaScript action for native Form Lock/Unlock button."""
        field_ids = []
        for spec in field_specs:
            if spec.field_id in ("Field_5_1", "GeneratedOutputText", "btnLockToggle", "btnGenerate", "btnReset"):
                continue
            if spec.field_datatype == 'multi_choice':
                raw_opts = spec.dropdown_options if spec.dropdown_options else ["Option 1", "Option 2"]
                opts = [o.strip() for o in raw_opts if o.strip()]
                for idx in range(len(opts)):
                    field_ids.append(f"{spec.field_id}_{idx}")
            else:
                field_ids.append(spec.field_id)

        if not field_ids:
            field_ids = ['Field_1_1', 'Field_1_2', 'Field_1_3']
        js_field_array = str(field_ids)

        js_code = (
            f"var fields = {js_field_array};\n"
            "var isCurrentlyLocked = false;\n"
            "var firstFld = this.getField(fields[0]);\n"
            "if (firstFld) { isCurrentlyLocked = firstFld.readonly; }\n"
            "for (var i = 0; i < fields.length; i++) {\n"
            "    var f = this.getField(fields[i]);\n"
            "    if (f) { f.readonly = !isCurrentlyLocked; }\n"
            "}\n"
            "if (isCurrentlyLocked) {\n"
            "    event.target.buttonSetCaption('[ UNLOCKED - LOCK EDITING ]');\n"
            "    event.target.fillColor = ['RGB', 0.12, 0.31, 0.47];\n"
            "    event.target.textColor = ['RGB', 1.0, 1.0, 1.0];\n"
            "} else {\n"
            "    event.target.buttonSetCaption('[ LOCKED - UNLOCK EDITING ]');\n"
            "    event.target.fillColor = ['RGB', 0.75, 0.0, 0.0];\n"
            "    event.target.textColor = ['RGB', 1.0, 1.0, 1.0];\n"
            "}\n"
        )
        return pdfdoc.PDFDictionary({
            'S': pdfdoc.PDFName('JavaScript'),
            'JS': pdfdoc.PDFString(js_code)
        })

    @staticmethod
    def build_summary_generator_action(field_specs: List['ExcelFieldSpec']) -> pdfdoc.PDFDictionary:
        """Returns dynamic Acrobat JavaScript action generating summary string from ALL user fields (including empty ones)."""
        js_lines = [
            "var summaryParts = [];",
            "var invalidFields = [];",
            "var f, val, displayVal, isBad;"
        ]

        for spec in field_specs:
            if spec.field_id in ("Field_5_1", "GeneratedOutputText", "btnLockToggle", "btnGenerate", "btnReset"):
                continue
            field_name = spec.field_id
            label = spec.field_label.strip().replace("'", "\\'")
            dtype = spec.field_datatype
            if not label:
                continue

            if dtype == 'multi_choice':
                raw_opts = spec.dropdown_options if spec.dropdown_options else ["Option 1", "Option 2"]
                opts = [o.strip() for o in raw_opts if o.strip()]
                opts_js_array = str(opts)
                js_lines.append(f"    var chkOpts_{field_name} = [];")
                js_lines.append(f"    var optsList_{field_name} = {opts_js_array};")
                js_lines.append(f"    for (var k = 0; k < optsList_{field_name}.length; k++) {{")
                js_lines.append(f"        var chk_{field_name} = this.getField('{field_name}_' + k);")
                js_lines.append(f"        if (chk_{field_name} && chk_{field_name}.value !== 'Off' && chk_{field_name}.value !== false && chk_{field_name}.value !== '') {{")
                js_lines.append(f"            chkOpts_{field_name}.push(optsList_{field_name}[k]);")
                js_lines.append(f"        }}")
                js_lines.append(f"    }}")
                js_lines.append(f"    displayVal = (chkOpts_{field_name}.length > 0) ? chkOpts_{field_name}.join(', ') : 'NIL';")
                js_lines.append(f"    summaryParts.push('{label}: ' + displayVal);")
            else:
                js_lines.append(f"    f = this.getField('{field_name}');")
                js_lines.append(f"    if (!f && '{field_name}' === 'Field_4_1') {{ f = this.getField('RawDprText'); }}")
                js_lines.append(f"    if (f) {{")
                js_lines.append(f"        val = (f.value !== undefined && f.value !== null) ? (f.value + '').replace(/[\\r\\n]+/g, ' ').trim() : '';")
                js_lines.append(f"        isBad = false;")

                if dtype == 'radio_choice':
                    js_lines.append(f"        val = (val !== '' && val !== 'Off') ? val : '';")
                    js_lines.append(f"        displayVal = (val !== '') ? val : 'NIL';")
                    js_lines.append(f"        summaryParts.push('{label}: ' + displayVal);")
                else:
                    if dtype == 'whole_number':
                        js_lines.append(f"        if (val !== '') {{ var n_{field_name} = Number(val); if (isNaN(n_{field_name}) || !Number.isInteger(n_{field_name})) isBad = true; }}")
                    elif dtype == 'decimal':
                        js_lines.append(f"        if (val !== '') {{ if (isNaN(Number(val))) isBad = true; }}")
                    elif dtype == 'date':
                        js_lines.append(f"        if (val !== '') {{ var p_{field_name} = val.split('-'); if (p_{field_name}.length !== 3 || isNaN(parseInt(p_{field_name}[0], 10)) || isNaN(parseInt(p_{field_name}[1], 10)) || isNaN(parseInt(p_{field_name}[2], 10))) isBad = true; }}")

                    js_lines.append(f"        if (isBad) {{")
                    js_lines.append(f"            f.fillColor = ['RGB', 1.0, 0.78, 0.81];")
                    js_lines.append(f"            f.textColor = ['RGB', 0.61, 0.0, 0.02];")
                    js_lines.append(f"            invalidFields.push('{label}');")
                    js_lines.append(f"        }} else {{")
                    js_lines.append(f"            displayVal = (val !== '' && val !== ' ') ? val : 'NIL';")
                    js_lines.append(f"            summaryParts.push('{label}: ' + displayVal);")
                    js_lines.append(f"        }}")
                js_lines.append(f"    }}")

        js_lines.append("var targetField = this.getField('GeneratedOutputText') || this.getField('Field_5_1');")
        js_lines.append("if (targetField) {")
        js_lines.append("    if (invalidFields.length > 0) {")
        js_lines.append("        var warnHeader = '• VALIDATION WARNING: ' + invalidFields.length + ' invalid field(s) detected [' + invalidFields.join(', ') + '] (highlighted in red). Correct them and click generate again.\\n\\n';")
        js_lines.append("        var restText = summaryParts.join(' | ');")
        js_lines.append("        try {")
        js_lines.append("            var sWarn = new Object(); sWarn.text = warnHeader; sWarn.textColor = color.red; sWarn.textFont = 'Helvetica-Bold'; sWarn.textSize = 9;")
        js_lines.append("            var sText = new Object(); sText.text = restText; sText.textColor = color.black; sText.textFont = 'Helvetica'; sText.textSize = 9;")
        js_lines.append("            targetField.richValue = [sWarn, sText];")
        js_lines.append("            targetField.richContents = [sWarn, sText];")
        js_lines.append("        } catch(eRich) {")
        js_lines.append("            targetField.value = warnHeader + restText;")
        js_lines.append("        }")
        js_lines.append("    } else {")
        js_lines.append("        var allText = summaryParts.join(' | ');")
        js_lines.append("        try {")
        js_lines.append("            var sText = new Object(); sText.text = allText; sText.textColor = color.black; sText.textFont = 'Helvetica'; sText.textSize = 9;")
        js_lines.append("            targetField.richValue = [sText];")
        js_lines.append("            targetField.richContents = [sText];")
        js_lines.append("        } catch(eRich) {")
        js_lines.append("            targetField.value = allText;")
        js_lines.append("        }")
        js_lines.append("    }")
        js_lines.append("}")

        js_code = "\n".join(js_lines)
        return pdfdoc.PDFDictionary({
            'S': pdfdoc.PDFName('JavaScript'),
            'JS': pdfdoc.PDFString(js_code)
        })

    @staticmethod
    def build_rugged_clear_action(field_specs: List['ExcelFieldSpec']) -> pdfdoc.PDFDictionary:
        """Returns dynamic Acrobat JavaScript action for clearing input fields without touching button background styling."""
        field_ids = []
        for spec in field_specs:
            if spec.field_id in ("btnLockToggle", "btnGenerate", "btnReset", "btnCopySummary"):
                continue
            if spec.field_datatype == 'multi_choice':
                raw_opts = spec.dropdown_options if spec.dropdown_options else ["Option 1", "Option 2"]
                opts = [o.strip() for o in raw_opts if o.strip()]
                for idx in range(len(opts)):
                    field_ids.append(f"{spec.field_id}_{idx}")
            else:
                field_ids.append(spec.field_id)

        js_field_array = str(field_ids)

        js_code = (
            f"var fields = {js_field_array};\n"
            "for (var i = 0; i < fields.length; i++) {\n"
            "    var f = this.getField(fields[i]);\n"
            "    if (f && !f.readonly) {\n"
            "        if (f.type === 'combobox') {\n"
            "            f.value = (f.numItems > 0) ? f.getItemAt(0) : ' ';\n"
            "        } else if (f.type === 'checkbox' || f.type === 'radiobutton') {\n"
            "            f.value = 'Off';\n"
            "        } else {\n"
            "            f.value = '';\n"
            "        }\n"
            "        if (f.name.toLowerCase().indexOf('auto') !== -1 && f.name.toLowerCase().indexOf('closing') !== -1) {\n"
            "            f.fillColor = ['RGB', 0.86, 0.92, 0.96];\n"
            "        } else {\n"
            "            f.fillColor = ['RGB', 0.98, 0.98, 0.98];\n"
            "        }\n"
            "        f.textColor = ['RGB', 0.0, 0.0, 0.0];\n"
            "    }\n"
            "}\n"
            "var outField = this.getField('GeneratedOutputText') || this.getField('Field_5_1');\n"
            "if (outField) { outField.value = ''; }\n"
        )
        return pdfdoc.PDFDictionary({
            'S': pdfdoc.PDFName('JavaScript'),
            'JS': pdfdoc.PDFString(js_code)
        })


def _stamp_page_numbers_if_multipage(pdf_path: str):
    """
    If PDF contains more than 1 page, stamps 'Page X of Y' in top right corner of every page.
    Preserves all AcroForm interactive fields and button JavaScript actions 100% intact.
    """
    try:
        reader = PdfReader(pdf_path)
        num_pages = len(reader.pages)
        if num_pages <= 1:
            return

        writer = PdfWriter()
        writer.append(reader)

        # Generate page number overlay PDF in memory
        packet = io.BytesIO()
        overlay_canvas = canvas.Canvas(packet, pagesize=letter)
        w, h = letter

        for p_num in range(1, num_pages + 1):
            overlay_canvas.setFillColor(HexColor("#FFFFFF"))
            overlay_canvas.setFont("Helvetica-Bold", 8.5)
            overlay_canvas.drawRightString(w - 25, h - 22, f"Page {p_num} of {num_pages}")
            overlay_canvas.showPage()
        overlay_canvas.save()

        packet.seek(0)
        overlay_reader = PdfReader(packet)

        for p_idx in range(num_pages):
            writer.pages[p_idx].merge_page(overlay_reader.pages[p_idx])

        with open(pdf_path, "wb") as f_out:
            writer.write(f_out)
    except Exception as e:
        logger.warning(f"Could not stamp page numbers: {e}")


def _center_pad_button_label(label: str, box_width: float, font_name: str = "Helvetica", font_size: float = 9.0) -> str:
    """
    LESSON LEARNED (Mathematical Space Padding Centering):
    ------------------------------------------------------
    Calculates exact leading space padding to visually center button labels inside ReportLab AcroForm textfields.
    This preserves ReportLab's pre-rendered '/AP' (Appearance Stream) vector dictionary so that background fill colors,
    borders, hand cursors, tooltips, and Acrobat JavaScript actions remain 100% visible and functional across all PDF viewers.
    """
    try:
        from reportlab.pdfbase.pdfmetrics import stringWidth
        txt_w = stringWidth(label.strip(), font_name, font_size)
        spc_w = stringWidth(" ", font_name, font_size)
        if box_width > txt_w and spc_w > 0:
            pad_spaces = int((box_width - txt_w) / (2.0 * spc_w))
            return (" " * pad_spaces) + label.strip()
    except Exception:
        pass
    return label


def _draw_page_header_banner(c, config, width, height):
    """Draws corporate Navy Primary header banner at top of canvas on every page."""
    c.setFillColor(HexColor(config.navy_primary))
    c.rect(0, height - 60, width, 60, fill=1, stroke=0)

    c.setFillColor(HexColor("#FFFFFF"))
    c.setFont("Helvetica-Bold", 14)
    banner_display_title = config.title.replace("_", " ") if "_" in config.title else config.title
    if config.subtitle.strip():
        c.drawCentredString(width / 2.0, height - 25, banner_display_title)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width / 2.0, height - 45, config.subtitle)
    else:
        c.drawCentredString(width / 2.0, height - 35, banner_display_title)


def generate_pdf_from_excel(
    excel_config_path: str = "Interactive_PDF_Code/fillable_pdf_configurator.xlsx",
    output_pdf_path: Optional[str] = None
) -> str:
    """
    Main PDF generation function. Reads configuration from master 3-sheet Excel workbook and
    compiles the interactive fillable PDF form.
    """
    config, section_specs, field_specs = ExcelConfigReader.read_config(excel_config_path)

    # Determine Output PDF Path relative to EXE/script base directory inside Interactive_PDF/
    if not output_pdf_path:
        base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(base_dir, "Interactive_PDF")
        stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        sanitized_title = config.title.replace(" ", "_")
        output_pdf_path = os.path.join(out_dir, f"{sanitized_title}_{stamp}.pdf")

    abs_pdf_path = os.path.abspath(output_pdf_path)
    os.makedirs(os.path.dirname(abs_pdf_path), exist_ok=True)

    c = canvas.Canvas(abs_pdf_path, pagesize=letter)
    width, height = config.page_width, config.page_height

    # 1. Render Header Banner on Page 1
    _draw_page_header_banner(c, config, width, height)

    # 2. Render Instructions Box (DIRECTIVE 10: OMIT ENTIRE BLOCK IF ALL INSTRUCTION POINTS ARE EMPTY!)
    mandatory_inst1 = "!! Save As - this PDF file before mailing. Cross-check whether all filled form data is properly saved before sending. !!"
    mandatory_inst2 = config.instruction_2.strip() if (config.instruction_2 and config.instruction_2.strip()) else "!! Before saving the file Click 'Generate Summary' Button to update it..!!   !! Use Adobe Acrobat Reader software for filling this form. !!"
    inst_points = [p.strip() for p in [mandatory_inst1, mandatory_inst2, config.instruction_3, config.instruction_4] if p.strip()]
    
    if inst_points:
        inst_height = 25 + (len(inst_points) * 14)
        c.setFillColor(HexColor(config.soft_blue))
        c.setStrokeColor(HexColor(config.navy_primary))
        c.setLineWidth(1)
        c.rect(30, height - 70 - inst_height, width - 60, inst_height, fill=1, stroke=1)

        c.setFillColor(HexColor(config.navy_primary))
        c.setFont("Helvetica-Bold", 9.5)
        c.drawString(40, height - 83, "INSTRUCTIONS FOR FILLING THIS FORM:")

        y_inst = height - 98
        for idx, p_text in enumerate(inst_points):
            txt_color = HexColor("#C00000") if idx in (0, 1) else HexColor(config.navy_primary)
            c.setFillColor(txt_color)
            c.setFont("Helvetica-Bold", 9.0)
            c.drawString(45, y_inst, p_text)
            y_inst -= 14

        y_curr = height - 80 - inst_height - 15
    else:
        y_curr = height - 75

    # Action Builders
    aa_num_fmt = AcrobatActionBuilder.build_decimal_validation_action()
    aa_whole_fmt = AcrobatActionBuilder.build_whole_number_validation_action()
    aa_date_fmt = AcrobatActionBuilder.build_date_validation_action()
    aa_closing_calc = AcrobatActionBuilder.build_closing_hsd_calc_action()
    js_lock_toggle = AcrobatActionBuilder.build_lock_toggle_action(field_specs)
    js_summary_gen = AcrobatActionBuilder.build_summary_generator_action(field_specs)
    js_rugged_clear = AcrobatActionBuilder.build_rugged_clear_action(field_specs)

    initial_readonly = 'readOnly' if (config.lock_status.strip().lower() == "locked") else ''
    show_lock_button = (config.lock_status.strip().lower() != "not required")

    lock_button_rendered = False

    # 3. DYNAMIC SECTION & FIELD RENDERER LOOP (PARSED FROM SHEET 2 & SHEET 3)
    if not section_specs:
        section_specs = [
            ExcelSectionSpec(1, "SEC_1", "1. GENERAL RIG & WELL IDENTIFICATION", 3),
            ExcelSectionSpec(2, "SEC_2", "2. HSD FUEL MANAGEMENT (KL)", 4),
            ExcelSectionSpec(3, "SEC_3", "3. WEATHER & LOGISTICS PARAMETERS", 3),
            ExcelSectionSpec(4, "SEC_4", "4. 24-HOUR OPERATIONAL PROGRESS NARRATIVE", 1)
        ]

    for sec_idx, sec in enumerate(section_specs):
        # Page Overflow Check: If remaining height is insufficient for section title, start a new page
        if y_curr < 100:
            if config.footer_text.strip():
                c.setFillColor(HexColor("#595959"))
                c.setFont("Helvetica-Oblique", 8)
                c.drawCentredString(width / 2.0, 25, config.footer_text)
            c.showPage()
            _draw_page_header_banner(c, config, width, height)
            y_curr = height - 75

        c.setFillColor(HexColor(config.navy_primary))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(30, y_curr, sec.section_title)

        if show_lock_button and not lock_button_rendered:
            is_initially_locked = (config.lock_status.strip().lower() == "locked")
            if is_initially_locked:
                btn_lock_label = "[ LOCKED - UNLOCK EDITING ]"
                btn_lock_bg = HexColor("#C00000")
                btn_lock_border = HexColor("#C00000")
            else:
                btn_lock_label = "[ UNLOCKED - LOCK EDITING ]"
                btn_lock_bg = HexColor(config.navy_primary)
                btn_lock_border = HexColor(config.navy_primary)

            btn_lock_display = _center_pad_button_label(btn_lock_label, 165, "Helvetica", 8.5)
            c.acroForm.textfield(
                name="btnLockToggle", value=btn_lock_display,
                x=width - 195, y=y_curr - 2, width=165, height=18,
                borderStyle='solid', borderWidth=1, borderColor=btn_lock_border, fillColor=btn_lock_bg,
                textColor=HexColor("#FFFFFF"), fontSize=8.5, tooltip="Click to lock/unlock form fields for editing", relative=True
            )
            btn_lock_ref = c.acroForm.fields[-1]
            btn_lock_obj = c._doc.idToObject[btn_lock_ref.name]
            btn_lock_obj.dict['FT'] = pdfdoc.PDFName('Btn')
            btn_lock_obj.dict['A'] = js_lock_toggle
            btn_lock_obj.dict['MK'] = pdfdoc.PDFDictionary({'CA': pdfdoc.PDFString(btn_lock_label)})
            btn_lock_obj.dict['H'] = pdfdoc.PDFName('P')  # Push highlight mode -> Pointing Hand Cursor in Acrobat
            btn_lock_obj.dict['Ff'] = 65536
            btn_lock_obj.dict['Q'] = 1
            lock_button_rendered = True

        c.setStrokeColor(HexColor("#B0C4DE"))
        c.setLineWidth(0.8)
        c.line(30, y_curr - 7, width - 30, y_curr - 7)

        sec_fields = [f for f in field_specs if f.section_id == sec.section_id]
        if not sec_fields:
            y_curr -= 35
            continue

        num_cols = max(1, min(4, sec.grid_cols))
        col_w = (width - 60 - (num_cols - 1) * 15) / float(num_cols)

        y_field_top = y_curr - 20
        for row_start in range(0, len(sec_fields), num_cols):
            row_fields = sec_fields[row_start:row_start + num_cols]
            has_long_text = any(f.field_datatype == 'long_text' for f in row_fields)
            field_height = 65 if has_long_text else 18

            # Page Overflow Check: If remaining height is insufficient for this field row, start a new page
            if y_field_top - (field_height + 25) < 45:
                if config.footer_text.strip():
                    c.setFillColor(HexColor("#595959"))
                    c.setFont("Helvetica-Oblique", 8)
                    c.drawCentredString(width / 2.0, 25, config.footer_text)
                c.showPage()
                _draw_page_header_banner(c, config, width, height)
                y_curr = height - 75
                y_field_top = y_curr - 20

            for col_idx, f_spec in enumerate(row_fields):
                x_pos = 30 + col_idx * (col_w + 15)
                c.setFillColor(HexColor("#595959"))
                c.setFont("Helvetica-Bold", 8.5)
                c.drawString(x_pos, y_field_top, f_spec.field_label)

                y_box = y_field_top - 20 - (field_height - 18)

                if f_spec.field_datatype == 'choice':
                    raw_opts = f_spec.dropdown_options if f_spec.dropdown_options else ["Option 1", "Option 2"]
                    opts = [o.strip() for o in raw_opts if o.strip()]
                    if " " not in opts:
                        opts = [" "] + opts
                    def_choice = f_spec.default_value.strip() if (f_spec.default_value and f_spec.default_value.strip() in opts) else " "
                    c.acroForm.choice(
                        name=f_spec.field_id, value=def_choice, options=opts,
                        x=x_pos, y=y_box, width=col_w, height=18,
                        borderStyle='solid', borderWidth=1, borderColor=HexColor("#B0C4DE"),
                        fillColor=HexColor("#FAFAFA"), textColor=HexColor("#000000"), fontSize=9.0,
                        fieldFlags='combo edit', tooltip=f_spec.tooltip or f_spec.field_label, relative=True
                    )
                    combo_flag = 131072 | 262144
                    if initial_readonly: combo_flag |= 1
                    c._doc.idToObject[c.acroForm.fields[-1].name].dict['Ff'] = combo_flag
                elif f_spec.field_datatype == 'multi_choice':
                    raw_opts = f_spec.dropdown_options if f_spec.dropdown_options else ["Option 1", "Option 2"]
                    opts = [o.strip() for o in raw_opts if o.strip()]
                    defaults = [d.strip() for d in f_spec.default_value.split(',') if d.strip()] if f_spec.default_value else []
                    opt_x = x_pos
                    opt_y = y_box + 2
                    for idx, opt in enumerate(opts):
                        sub_name = f"{f_spec.field_id}_{idx}"
                        is_checked = (opt in defaults)
                        c.acroForm.checkbox(
                            name=sub_name, checked=is_checked, size=14,
                            x=opt_x, y=opt_y, buttonStyle='check',
                            shape='square', fillColor=HexColor("#FAFAFA"), borderColor=HexColor("#B0C4DE"),
                            borderStyle='solid', borderWidth=1, relative=True
                        )
                        if initial_readonly == 'readOnly':
                            try:
                                c._doc.idToObject[c.acroForm.fields[-1].name].dict['Ff'] = 1
                            except Exception:
                                pass
                        c.setFillColor(HexColor("#000000"))
                        c.setFont("Helvetica-Bold", 8.5)
                        c.drawString(opt_x + 18, opt_y + 3, opt)
                        opt_x += (len(opt) * 6 + 34)
                        if opt_x > x_pos + col_w - 40:
                            opt_x = x_pos
                            opt_y -= 16
                elif f_spec.field_datatype == 'radio_choice':
                    raw_opts = f_spec.dropdown_options if f_spec.dropdown_options else ["Yes", "No"]
                    opts = [o.strip() for o in raw_opts if o.strip()]
                    if len(opts) < 2:
                        opts = ["Yes", "No"]
                    def_val = f_spec.default_value.strip() if f_spec.default_value else ""
                    opt_x = x_pos
                    opt_y = y_box + 2
                    for opt in opts:
                        is_selected = (opt == def_val)
                        c.acroForm.radio(
                            name=f_spec.field_id, value=opt,
                            selected=is_selected, size=14,
                            x=opt_x, y=opt_y, buttonStyle='circle',
                            shape='circle', fillColor=HexColor("#FAFAFA"), borderColor=HexColor("#B0C4DE"),
                            borderStyle='solid', borderWidth=1, relative=True
                        )
                        c.setFillColor(HexColor("#000000"))
                        c.setFont("Helvetica-Bold", 8.5)
                        c.drawString(opt_x + 18, opt_y + 3, opt)
                        opt_x += (len(opt) * 6 + 34)
                        if opt_x > x_pos + col_w - 40:
                            opt_x = x_pos
                            opt_y -= 16
                    if initial_readonly == 'readOnly':
                        try:
                            rg_obj = c._doc.idToObject[c.acroForm.fields[-1].name]
                            if hasattr(rg_obj, 'Ff'):
                                rg_obj.Ff |= 1
                        except Exception:
                            pass
                elif f_spec.field_datatype == 'long_text':
                    multi_w = (width - 60) if num_cols == 1 else col_w
                    multi_flags = 'multiline ' + initial_readonly if initial_readonly else 'multiline'
                    c.acroForm.textfield(
                        name=f_spec.field_id, value=f_spec.default_value,
                        x=x_pos, y=y_box, width=multi_w, height=field_height,
                        borderStyle='solid', borderWidth=1, borderColor=HexColor("#B0C4DE"),
                        fillColor=HexColor("#FAFAFA"), textColor=HexColor("#000000"), fontSize=9.0,
                        maxlen=4000, fieldFlags=multi_flags.strip(), tooltip=f_spec.tooltip or f_spec.field_label, relative=True
                    )
                else:
                    is_auto_closing = ("auto" in f_spec.field_label.lower() and "closing" in f_spec.field_label.lower())
                    box_fill = HexColor(config.soft_blue) if is_auto_closing else HexColor("#FAFAFA")
                    box_border = HexColor(config.navy_primary) if is_auto_closing else HexColor("#B0C4DE")
                    box_text_color = HexColor(config.navy_primary) if is_auto_closing else HexColor("#000000")
                    f_flags = 'readOnly' if is_auto_closing else initial_readonly

                    c.acroForm.textfield(
                        name=f_spec.field_id, value=f_spec.default_value,
                        x=x_pos, y=y_box, width=col_w, height=18,
                        borderStyle='solid', borderWidth=1.5 if is_auto_closing else 1,
                        borderColor=box_border, fillColor=box_fill, textColor=box_text_color, fontSize=9.0,
                        maxlen=200, fieldFlags=f_flags, tooltip=f_spec.tooltip or f_spec.field_label, relative=True
                    )
                    if is_auto_closing:
                        c._doc.idToObject[c.acroForm.fields[-1].name].dict['AA'] = aa_closing_calc
                    elif f_spec.field_datatype == 'whole_number':
                        c._doc.idToObject[c.acroForm.fields[-1].name].dict['AA'] = aa_whole_fmt
                    elif f_spec.field_datatype == 'decimal':
                        c._doc.idToObject[c.acroForm.fields[-1].name].dict['AA'] = aa_num_fmt
                    elif f_spec.field_datatype == 'date':
                        c._doc.idToObject[c.acroForm.fields[-1].name].dict['AA'] = aa_date_fmt

            y_field_top -= (field_height + 25)

        y_curr = y_field_top - 15

    # 4. RENDER SUMMARY GENERATOR & RESET BUTTONS (CONFIGURED FROM SHEET 1)
    # If remaining canvas height is insufficient for summary block, start a new page
    if y_curr < 180:
        c.showPage()
        y_curr = height - 50

    y_section_5 = y_curr
    c.setFillColor(HexColor(config.navy_primary))
    c.setFont("Helvetica-Bold", 10)
    sum_sec_title = config.summary_section_title or "GENERATE FORM SUMMARY & COPY TEXT"
    c.drawString(30, y_section_5, sum_sec_title)
    c.setStrokeColor(HexColor("#B0C4DE"))
    c.setLineWidth(0.8)
    c.line(30, y_section_5 - 4, width - 30, y_section_5 - 4)

    btn_gen_label = config.summary_button_text or ">>>> Click the Button - To Generate Form Summary <<<<"
    btn_gen_width = width - 60 - 135
    btn_gen_display = _center_pad_button_label(btn_gen_label, btn_gen_width, "Helvetica-Bold", 9.5)

    c.acroForm.textfield(
        name="btnGenerate", value=btn_gen_display,
        x=30, y=y_section_5 - 38, width=btn_gen_width, height=28,
        borderStyle='solid', borderWidth=1.5, borderColor=HexColor("#C00000"), fillColor=HexColor(config.soft_yellow),
        textColor=HexColor("#C00000"), fontSize=9.5, tooltip="Click to update the summary text string below", relative=True
    )
    btn_ref = c.acroForm.fields[-1]
    btn_obj = c._doc.idToObject[btn_ref.name]
    btn_obj.dict['FT'] = pdfdoc.PDFName('Btn')
    btn_obj.dict['A'] = js_summary_gen
    btn_obj.dict['MK'] = pdfdoc.PDFDictionary({'CA': pdfdoc.PDFString(btn_gen_label)})
    btn_obj.dict['H'] = pdfdoc.PDFName('P')  # Push highlight mode -> Pointing Hand Cursor in Acrobat
    btn_obj.dict['Ff'] = 65536
    btn_obj.dict['Q'] = 1

    btn_rst_label = "[ CLEAR / RESET FORM ]"
    btn_rst_display = _center_pad_button_label(btn_rst_label, 125, "Helvetica-Bold", 9.0)
    x_reset = 30 + btn_gen_width + 10

    c.acroForm.textfield(
        name="btnReset", value=btn_rst_display,
        x=x_reset, y=y_section_5 - 38, width=125, height=28,
        borderStyle='solid', borderWidth=1.5, borderColor=HexColor(config.navy_primary), fillColor=HexColor("#E2E2E2"),
        textColor=HexColor(config.navy_primary), fontSize=9.0, tooltip="Click to clear all input fields and reset form to default values", relative=True
    )
    btn_rst_ref = c.acroForm.fields[-1]
    btn_rst_obj = c._doc.idToObject[btn_rst_ref.name]
    btn_rst_obj.dict['FT'] = pdfdoc.PDFName('Btn')
    btn_rst_obj.dict['A'] = js_rugged_clear
    btn_rst_obj.dict['MK'] = pdfdoc.PDFDictionary({'CA': pdfdoc.PDFString(btn_rst_label)})
    btn_rst_obj.dict['H'] = pdfdoc.PDFName('P')  # Push highlight mode -> Pointing Hand Cursor in Acrobat
    btn_rst_obj.dict['Ff'] = 65536
    btn_rst_obj.dict['Q'] = 1

    sum_fld_label = config.summary_field_label or "GENERATED FORM SUMMARY TEXT"
    c.acroForm.textfield(
        name="GeneratedOutputText",
        value="",
        x=30, y=y_section_5 - 135, width=width - 60, height=85,
        borderStyle='solid', borderWidth=1, borderColor=HexColor(config.navy_primary), fillColor=HexColor(config.soft_yellow),
        textColor=HexColor(config.navy_primary), fontSize=9.0, maxlen=8000, fieldFlags='multiline richText',
        tooltip=sum_fld_label, relative=True
    )

    out_ref = c.acroForm.fields[-1]
    out_obj = c._doc.idToObject[out_ref.name]
    out_obj.dict['Ff'] = 1 | 4096 | 0x2000000  # ReadOnly (1) + Multiline (4096) + RichText (0x2000000)
    out_obj.dict['DS'] = pdfdoc.PDFString('font: Helvetica 9pt; text-align: left; color: #000000;')
    out_obj.dict['RV'] = pdfdoc.PDFString('<?xml version="1.0"?><body xmlns="http://www.w3.org/1999/xhtml" style="font-size:9pt;font-family:Helvetica;color:#000000;"><p></p></body>')

    # 5. FOOTER BANNER
    if config.footer_text.strip():
        c.setFillColor(HexColor("#595959"))
        c.setFont("Helvetica-Oblique", 8)
        c.drawCentredString(width / 2.0, 25, config.footer_text)

    c.showPage()
    c.save()
    _stamp_page_numbers_if_multipage(abs_pdf_path)
    logger.info(f"Successfully generated Excel-driven fillable PDF form at: '{abs_pdf_path}'")
    return abs_pdf_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    excel_path = "Interactive_PDF_Code/fillable_pdf_configurator.xlsx"
    path = generate_pdf_from_excel(excel_path)
    print("Master Excel-Driven Fillable PDF Form Created at:", path)
