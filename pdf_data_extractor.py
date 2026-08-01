"""
PDF AcroForm Data Extraction & Excel Consolidation Engine
==========================================================
Scans a designated folder for completed interactive fillable PDF form files, extracts native
AcroForm field key-value pairs using 'pypdf', and generates a formatted corporate Excel
consolidation report:
'Interactive_PDF_Code/Interactive_PDF/Consolidate_Data_{YYYYMMDD_HHMMSS}.xlsx'

Column Structure:
-----------------
- Column A: Source_PDF_Filename
- Column B: RigName (Field_1_1)
- Column C: WellId (Field_1_2)
- Column D: ReportDate (Field_1_3)
- Column E: Stock_HSD_KL (Field_2_1)
- Column F: Requirement_HSD_KL (Field_2_2)
- Column G: Consumption_HSD_KL (Field_2_3)
- Column H: Closing_HSD_KL (Field_2_4)
- Column I: Weather_SwellHeight_Meters (Field_3_1)
- Column J: IsCurrentlyWaiting (Field_3_2)
- Column K: FoodBox_LastDate (Field_3_3)
- Column L: RawDprText (Field_4_1)
- Column M: GeneratedOutputText (Field_5_1)

Follows project coding standards (Python 3.10+ types, openpyxl, pypdf, logging).
"""

import os
import glob
import logging
import datetime
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

logger = logging.getLogger(__name__)


def extract_acroform_data_from_pdf(pdf_path: str) -> Dict[str, str]:
    """
    Extracts native AcroForm field values from a single PDF file using pypdf.

    Args:
        pdf_path (str): File path to target fillable PDF form.

    Returns:
        Dict[str, str]: Dictionary mapping AcroForm field names to extracted string values.
    """
    field_data: Dict[str, str] = {}
    try:
        reader = PdfReader(pdf_path)
        fields = reader.get_fields()
        if fields:
            for f_key, f_obj in fields.items():
                val = f_obj.get('/V', '')
                # Clean up PDF string formatting if necessary
                if isinstance(val, str):
                    field_data[f_key] = val.strip()
                elif val is not None:
                    field_data[f_key] = str(val).strip()
                else:
                    field_data[f_key] = ''
    except Exception as e:
        logger.warning(f"Could not extract AcroForm fields from '{pdf_path}': {e}")

    return field_data


def consolidate_pdf_forms_to_excel(
    pdf_folder_path: str = "Interactive_PDF_Code/Interactive_PDF",
    output_excel_path: Optional[str] = None
) -> Tuple[str, int]:
    """
    Scans specified folder for PDF files, extracts AcroForm field data from all PDFs,
    and compiles a formatted corporate consolidation Excel report.

    Args:
        pdf_folder_path (str): Target directory containing filled PDF forms.
        output_excel_path (Optional[str]): Path for output Excel report. If None, auto-generates timestamped filename.

    Returns:
        Tuple[str, int]: Absolute path to created Excel report and count of processed PDF files.
    """
    abs_folder = os.path.abspath(pdf_folder_path)
    os.makedirs(abs_folder, exist_ok=True)  # Auto-create folder if it does not exist

    pdf_files = glob.glob(os.path.join(abs_folder, "*.pdf"))
    if not pdf_files:
        logger.warning(f"No PDF files found in directory: '{abs_folder}'")

    # Define Consolidation Columns
    headers = [
        "Source_PDF_Filename",
        "RigName",
        "WellId",
        "ReportDate",
        "Stock_HSD_KL",
        "Requirement_HSD_KL",
        "Consumption_HSD_KL",
        "Closing_HSD_KL",
        "Weather_SwellHeight_Meters",
        "IsCurrentlyWaiting",
        "FoodBox_LastDate",
        "RawDprText",
        "GeneratedOutputText"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidated_PDF_Data"
    ws.views.sheetView[0].showGridLines = True

    # Styling Tokens
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_bold_navy = Font(name="Calibri", size=10, bold=True, color="1F4E78")
    font_regular = Font(name="Calibri", size=10, color="000000")
    
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border_thin = Border(
        left=Side(style='thin', color='B0C4DE'),
        right=Side(style='thin', color='B0C4DE'),
        top=Side(style='thin', color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE')
    )

    # Append Header Row
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_thin

    # Process PDF files row by row
    processed_count = 0
    for pdf_path in pdf_files:
        fname = os.path.basename(pdf_path)
        data = extract_acroform_data_from_pdf(pdf_path)

        row_vals = [
            fname,
            data.get('RigName', ''),
            data.get('WellId', ''),
            data.get('ReportDate', ''),
            data.get('Stock_HSD_KL', ''),
            data.get('Requirement_HSD_KL', ''),
            data.get('Consumption_HSD_KL', ''),
            data.get('Closing_HSD_KL', ''),
            data.get('Weather_SwellHeight_Meters', ''),
            data.get('IsCurrentlyWaiting', ''),
            data.get('FoodBox_LastDate', ''),
            data.get('RawDprText', ''),
            data.get('GeneratedOutputText', '')
        ]

        ws.append(row_vals)
        processed_count += 1

        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_bold_navy if col_idx == 1 else font_regular
            cell.border = border_thin
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
            if col_idx in (3, 4, 10, 11):
                cell.alignment = align_center

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 50), 16)

    # Determine Output Report Path
    if not output_excel_path:
        stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        output_excel_path = os.path.join(abs_folder, f"Consolidate_Data_{stamp}.xlsx")

    abs_out_path = os.path.abspath(output_excel_path)
    os.makedirs(os.path.dirname(abs_out_path), exist_ok=True)

    # Save Excel report (Handle PermissionError if open)
    try:
        wb.save(abs_out_path)
        logger.info(f"Successfully generated Consolidated Excel Report ({processed_count} PDFs) at: '{abs_out_path}'")
    except PermissionError:
        raise PermissionError(f"File Permission Error: '{abs_out_path}' is currently open in Microsoft Excel. Please close it and retry.")

    return abs_out_path, processed_count


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    out_file, count = consolidate_pdf_forms_to_excel("Interactive_PDF_Code/Interactive_PDF")
    print(f"Consolidated {count} PDF forms into Excel report: {out_file}")
