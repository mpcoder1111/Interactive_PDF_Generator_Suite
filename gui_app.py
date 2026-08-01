"""
Interactive Fillable PDF Form Suite — Standalone Tkinter GUI Configurator
===========================================================================
Complete, native Tkinter Graphical User Interface (GUI) Configurator providing:
1. Sheet 1 Editor: Global Form Basic Details (Title, Subtitle, Instructions, Lock Status, Colors).
2. Sheet 2 Editor: Section Details & Layout Grid Columns (Section ID Locked, Title & Grid Columns Editable).
3. Sheet 3 Editor: Field Configuration Schema (Field_ID & Section_ID Protected, Choice Option Validation, Preserved Row Selection).
4. Refined Modern Typography & Medium Blue Buttons: Medium-sized navy blue buttons, clean tab headers, and tab hover tooltips.
5. Fillable PDF Form Compiler: Generates interactive PDF forms directly from native Tkinter UI state.
6. Batch PDF Consolidator: Extracts AcroForm field data from PDFs into consolidated Excel reports.
7. Zero Excel References in UI: Clean, standalone Tkinter frontend interface.
8. Zero Modal Popups: Inline live status feedback bar.
9. Session File Logging: Writes diagnostics to 'Interactive_PDF_Code/logs/log_{YYYYMMDD_HHMMSS}.txt'.

Author: Interactive PDF Form Compiler Pipeline
Standards Spec: standards/excel_form_configurator_contracts.md
"""

import os
import sys
import time
import json
import logging
import datetime
import traceback
import webbrowser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
import tkinter as tk
from tkinter import ttk, filedialog

# Import backend modules from Interactive_PDF_Code package
try:
    from build_pdf_configurator_excel import create_3sheet_pdf_configurator_excel, validate_excel_config
    from pdf_generator_from_excel import generate_pdf_from_excel
    from pdf_data_extractor import consolidate_pdf_forms_to_excel
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from build_pdf_configurator_excel import create_3sheet_pdf_configurator_excel, validate_excel_config
    from pdf_generator_from_excel import generate_pdf_from_excel
    from pdf_data_extractor import consolidate_pdf_forms_to_excel

def get_base_dir():
    """Returns directory path where executable or script resides."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


# Initialize File Logging (log_{YYYYMMDD_HHMMSS}.txt)
def setup_file_logger():
    base_dir = get_base_dir()
    log_dir = os.path.join(base_dir, "logs")
    if os.path.isfile(log_dir):
        try:
            os.remove(log_dir)
        except Exception:
            pass
    os.makedirs(log_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir, f"log_{stamp}.txt")

    logger = logging.getLogger("DigitalDPRSuite")
    logger.setLevel(logging.INFO)

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    return logger, log_file


logger, LOG_FILE_PATH = setup_file_logger()


class ToolTip:
    """Hover Tooltip helper for Tkinter widgets and Notebook tabs."""

    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self.widget.bind("<Enter>", self.show_tip)
        self.widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tip_window or not self.text:
            return
        x, y, _, cy = self.widget.bbox("insert") if hasattr(self.widget, "bbox") else (0, 0, 0, 0)
        x = x + self.widget.winfo_rootx() + 25
        y = y + cy + self.widget.winfo_rooty() + 25

        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")

        label = tk.Label(
            tw, text=self.text, justify="left",
            bg="#1E1E1E", fg="#FFFFFF", relief="solid", bd=1,
            font=("Segoe UI", 9), padx=8, pady=4
        )
        label.pack(ipadx=1)

    def hide_tip(self, event=None):
        tw = self.tip_window
        self.tip_window = None
        if tw:
            tw.destroy()


class DigitalDPRSuiteGUI(tk.Tk):
    """
    Complete Standalone Tkinter GUI Form Configurator for Digital DPR.
    """

    def __init__(self):
        super().__init__()
        self.title("Interactive Fillable PDF Form Suite")
        self.geometry("1080x810")
        self.minsize(980, 700)
        self.configure(bg="#F4F6F9")

        base_dir = get_base_dir()
        self.excel_path = os.path.join(base_dir, "fillable_pdf_configurator.xlsx")
        self.output_pdf_folder = os.path.join(base_dir, "Interactive_PDF")
        os.makedirs(self.output_pdf_folder, exist_ok=True)

        # Configure Global ttk Styles (Enhanced Fonts & Row Heights)
        self._configure_ttk_styles()

        # --- 1. SHEET 1: Global Form Settings State ---
        self.form_settings = {
            "Form_Title": "Employee_Information_Form",
            "Form_Subtitle": "OFFICIAL EMPLOYEE DETAILS & RECORD FORM",
            "Instruction_Point_1": "!! Save As - this PDF file before mailing. Cross-check whether all filled form data is properly saved before sending. !!",
            "Instruction_Point_2": "!! Before saving the file Click 'Generate Summary' Button to update it..!!   !! Use Adobe Acrobat Reader software for filling this form. !!",
            "Instruction_Point_3": "3. Verify all employee personal details & department info before generating summary.",
            "Instruction_Point_4": "",
            "Footer_Text": "Official Employee Information & Department Record Form",
            "Default_Lock_Status": "Unlocked",
            "Theme_Navy_Primary": "#1F4E78",
            "Theme_Soft_Blue": "#DDEBF7",
            "Theme_Soft_Yellow": "#FFF2CC",
            "Summary_Section_Title": "GENERATE FORM SUMMARY & COPY TEXT",
            "Summary_Button_Text": ">>>> Click the Button - To Generate Form Summary <<<<",
            "Summary_Field_Label": "GENERATED FORM SUMMARY TEXT"
        }

        # --- 2. SHEET 2: Section Details State (User Form Layout Sections) ---
        self.section_rows = [
            [1, "SEC_1", "1. EMPLOYEE PERSONAL DETAILS", "3"],
            [2, "SEC_2", "2. DEPARTMENT & POSITION DETAILS", "3"],
            [3, "SEC_3", "3. RESIDENTIAL ADDRESS & REMARKS", "1"]
        ]

        # --- 3. SHEET 3: Field Configuration Schema State (User Form Fields) ---
        self.field_schema_rows = [
            ["Field_1_1", "SEC_1", "FULL NAME", "short_text", "", "Enter employee full legal name", "max_length", "100", ""],
            ["Field_1_2", "SEC_1", "EMPLOYEE ID / BADGE NO", "short_text", "", "Enter official Employee ID or Badge number", "max_length", "50", ""],
            ["Field_1_3", "SEC_1", "DATE OF BIRTH (DD-MM-YYYY)", "date", "", "Enter date of birth in DD-MM-YYYY format", "date_ddmmyyyy", "", ""],
            ["Field_1_4", "SEC_1", "CONTACT PHONE NUMBER", "whole_number", "", "Enter 10-digit mobile contact number", ">=", "0", ""],
            ["Field_1_5", "SEC_1", "OFFICIAL EMAIL ADDRESS", "short_text", "", "Enter company official email address", "max_length", "100", ""],
            ["Field_2_1", "SEC_2", "DEPARTMENT NAME", "choice", "", "Select employee department", "none", "", "Operations, Engineering, Human Resources, Finance, IT & Systems, Sales & Marketing, Executive Management"],
            ["Field_2_2", "SEC_2", "DESIGNATION / POSITION", "short_text", "", "Enter job designation or position title", "max_length", "100", ""],
            ["Field_2_3", "SEC_2", "JOINING DATE (DD-MM-YYYY)", "date", "", "Enter company joining date in DD-MM-YYYY format", "date_ddmmyyyy", "", ""],
            ["Field_2_4", "SEC_2", "MONTHLY BASIC SALARY ($)", "decimal", "", "Enter monthly basic salary in USD", ">=", "0.0", ""],
            ["Field_3_1", "SEC_3", "RESIDENTIAL ADDRESS & REMARKS", "long_text", "", "Enter complete residential address and additional notes", "min_length", "10", ""],
            ["Field_3_2", "SEC_3", "EMERGENCY CONTACT & PHONE", "short_text", "", "Enter emergency contact person name and telephone number", "max_length", "150", ""]
        ]

        # Dependent Operators Map for Pure Data Types
        self.operator_matrix = {
            "short_text": ["none", "min_length", "max_length"],
            "long_text": ["none", "min_length", "max_length"],
            "whole_number": ["none", "=", "<", ">", "<=", ">=", "in_range", "out_range"],
            "decimal": ["none", "=", "<", ">", "<=", ">=", "in_range", "out_range"],
            "choice": ["none"],
            "multi_choice": ["none"],
            "radio_choice": ["none"],
            "date": ["none", "date_ddmmyyyy", "=", "<", ">", "<=", ">=", "in_range", "out_range"]
        }

        self._build_ui()
        self.load_initial_state_from_excel()
        self.log_info(f"Tkinter Application Initialized. Logging to: '{LOG_FILE_PATH}'")

    def _configure_ttk_styles(self):
        """Sets crisp, readable Segoe UI typography and comfortable row heights for ttk widgets."""
        style = ttk.Style()
        style.theme_use("clam")

        # Notebook Tab Font
        style.configure("TNotebook.Tab", font=("Segoe UI", 10, "bold"), padding=[10, 5])
        
        # Treeview Formatting
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#E2E2E2", foreground="#1F4E78", padding=5)
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=26, background="#FFFFFF", fieldbackground="#FFFFFF")
        style.map("Treeview", background=[("selected", "#1F4E78")], foreground=[("selected", "#FFFFFF")])

    def _build_ui(self):
        """Constructs the Tkinter UI layout without Excel references or popups."""
        # 1. Header Banner
        header_frame = tk.Frame(self, bg="#1F4E78", height=70)
        header_frame.pack(fill="x", side="top")

        # Right-side Help Guide Button
        btn_help = tk.Button(
            header_frame,
            text="📖 User Guide (HTML)",
            font=("Segoe UI", 9, "bold"),
            bg="#0066CC",
            fg="#FFFFFF",
            activebackground="#004C99",
            activeforeground="#FFFFFF",
            bd=0,
            padx=12,
            pady=5,
            cursor="hand2",
            command=self.open_user_guide_in_browser
        )
        btn_help.pack(side="right", padx=15, pady=15)

        title_sub_frame = tk.Frame(header_frame, bg="#1F4E78")
        title_sub_frame.pack(side="left", padx=15, pady=8)

        lbl_title = tk.Label(
            title_sub_frame,
            text="INTERACTIVE FILLABLE PDF FORM SUITE",
            font=("Segoe UI", 15, "bold"),
            fg="#FFFFFF",
            bg="#1F4E78"
        )
        lbl_title.pack(anchor="w")

        lbl_subtitle = tk.Label(
            title_sub_frame,
            text="Domain-Agnostic Fillable PDF Form Generator & AcroForm Data Consolidator",
            font=("Segoe UI", 9, "italic"),
            fg="#DDEBF7",
            bg="#1F4E78"
        )
        lbl_subtitle.pack(anchor="w")

        # 2. Global Inline Live Status Feedback Bar (Top Level Status Bar with 2-Line Auto-Wrap)
        self.status_frame = tk.Frame(self, bg="#DDEBF7", bd=1, relief="solid")
        self.status_frame.pack(fill="x", side="top")

        self.lbl_status = tk.Label(
            self.status_frame,
            text="[STATUS] Ready. Configure your form settings below and click Create Fillable PDF Form.",
            font=("Segoe UI", 9, "bold"),
            fg="#1F4E78",
            bg="#DDEBF7",
            anchor="w",
            justify="left",
            padx=15,
            pady=5
        )
        self.lbl_status.pack(fill="x", expand=True, side="left")

        def _on_status_frame_resize(event):
            if event.width > 50:
                self.lbl_status.configure(wraplength=max(event.width - 30, 200))

        self.status_frame.bind("<Configure>", _on_status_frame_resize)

        # 3. Primary Compile & Configuration Action Toolbar (Top Level)
        action_bar = tk.Frame(self, bg="#F4F6F9", pady=8, padx=15)
        action_bar.pack(fill="x", side="top")

        btn_compile = tk.Button(
            action_bar,
            text="⚡ Create Fillable PDF Form",
            font=("Segoe UI", 10, "bold"),
            bg="#1F4E78",
            fg="#FFFFFF",
            activebackground="#143451",
            activeforeground="#FFFFFF",
            relief="solid",
            bd=1,
            padx=12,
            pady=6,
            cursor="hand2",
            command=self.compile_fillable_pdf
        )
        btn_compile.pack(side="left", padx=(0, 4))

        btn_export_excel = tk.Button(
            action_bar,
            text="💾 Export Excel",
            font=("Segoe UI", 10, "bold"),
            bg="#E2E2E2",
            fg="#1F4E78",
            activebackground="#D0D0D0",
            activeforeground="#1F4E78",
            relief="solid",
            bd=1,
            padx=10,
            pady=6,
            cursor="hand2",
            command=self.export_excel_dialog
        )
        btn_export_excel.pack(side="left", padx=3)

        btn_export_json = tk.Button(
            action_bar,
            text="📄 Export JSON",
            font=("Segoe UI", 10, "bold"),
            bg="#E2E2E2",
            fg="#1F4E78",
            activebackground="#D0D0D0",
            activeforeground="#1F4E78",
            relief="solid",
            bd=1,
            padx=10,
            pady=6,
            cursor="hand2",
            command=self.export_json_dialog
        )
        btn_export_json.pack(side="left", padx=3)

        btn_ai_schema = tk.Button(
            action_bar,
            text="📋 Download AI Schema",
            font=("Segoe UI", 10, "bold"),
            bg="#FFF2CC",
            fg="#7F6000",
            activebackground="#FFE699",
            activeforeground="#7F6000",
            relief="solid",
            bd=1,
            padx=10,
            pady=6,
            cursor="hand2",
            command=self.download_ai_schema_dialog
        )
        btn_ai_schema.pack(side="left", padx=3)

        btn_import_cfg = tk.Button(
            action_bar,
            text="📂 Import Config",
            font=("Segoe UI", 10, "bold"),
            bg="#E2E2E2",
            fg="#1F4E78",
            activebackground="#D0D0D0",
            activeforeground="#1F4E78",
            relief="solid",
            bd=1,
            padx=10,
            pady=6,
            cursor="hand2",
            command=self.import_config_dialog
        )
        btn_import_cfg.pack(side="left", padx=3)

        # 4. Tabbed Configuration Notebook
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        tab0 = tk.Frame(self.notebook, bg="#F4F6F9")
        self.notebook.add(tab0, text=" ℹ️ About ")

        tab1 = tk.Frame(self.notebook, bg="#F4F6F9")
        self.notebook.add(tab1, text=" 📋 Sheet 1: Basic Details ")

        tab2 = tk.Frame(self.notebook, bg="#F4F6F9")
        self.notebook.add(tab2, text=" 📑 Sheet 2: Section Layouts ")

        tab3 = tk.Frame(self.notebook, bg="#F4F6F9")
        self.notebook.add(tab3, text=" ⚙️ Sheet 3: Field Schema ")

        tab4 = tk.Frame(self.notebook, bg="#F4F6F9")
        self.notebook.add(tab4, text=" 📊 Batch PDF Consolidator ")

        tab5 = tk.Frame(self.notebook, bg="#F4F6F9")
        self.notebook.add(tab5, text=" 📄 Session Logs ")

        # Bind Mouse Motion for Full Tab Hover Tooltips
        self.tab_tooltips = {
            0: "About Software Suite: Overview of core capabilities & key application uses",
            1: "Sheet 1: Form Basic Details (Title, Subtitle, Instructions & Lock Status)",
            2: "Sheet 2: Section Details Layout Architecture & Grid Column Allocations",
            3: "Sheet 3: Field Configuration Schema (Datatypes, Operators, Dropdowns)",
            4: "Batch PDF Data Extraction & Consolidated Excel Report Export",
            5: "Session Log Output Console & File Diagnostics"
        }
        self.notebook.bind("<Motion>", self.on_notebook_hover)

        self._build_tab0_about(tab0)
        self._build_tab1_basic_details(tab1)
        self._build_tab2_sections(tab2)
        self._build_tab3_fields(tab3)
        self._build_tab4_consolidator(tab4)
        self._build_tab5_logs(tab5)

    def on_notebook_hover(self, event):
        """Displays full tab title on mouse hover if cursor is over a notebook tab."""
        try:
            tab_index = self.notebook.index(f"@{event.x},{event.y}")
            tip_text = self.tab_tooltips.get(tab_index, "")
            if tip_text:
                # Update status bar quietly on hover
                pass
        except tk.TclError:
            pass

    def set_inline_status(self, text: str, is_error: bool = False, is_success: bool = False):
        """Updates inline status bar text and background color cleanly without popups."""
        bg_color = "#DDEBF7"
        fg_color = "#1F4E78"

        if is_error:
            bg_color = "#FFC7CE"  # Soft Red
            fg_color = "#C00000"
        elif is_success:
            bg_color = "#C6EFCE"  # Soft Green
            fg_color = "#006100"

        self.status_frame.configure(bg=bg_color)
        self.lbl_status.configure(text=f"[STATUS] {text}", bg=bg_color, fg=fg_color)
        self.update_idletasks()

    def log_info(self, message: str):
        """Logs info message to console text widget and log file."""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        formatted = f"[{timestamp}] [INFO] {message}\n"
        if hasattr(self, 'txt_console'):
            self.txt_console.insert(tk.END, formatted)
            self.txt_console.see(tk.END)
        logger.info(message)

    def log_error(self, message: str):
        """Logs error message to console text widget and log file."""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        formatted = f"[{timestamp}] [ERROR] {message}\n"
        if hasattr(self, 'txt_console'):
            self.txt_console.insert(tk.END, formatted)
            self.txt_console.see(tk.END)
        logger.error(message)

    # =========================================================================
    # TAB 0: About Software Overview
    # =========================================================================
    def _build_tab0_about(self, parent):
        """Builds the About software tab providing an intuitive, professional overview of core features."""
        canvas = tk.Canvas(parent, bg="#F4F6F9", highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="#F4F6F9", padx=25, pady=20)

        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=v_scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")

        # 1. HERO HEADER BANNER
        header_card = tk.Frame(scroll_frame, bg="#1F4E78", padx=20, pady=15, relief="solid", bd=1)
        header_card.pack(fill="x", expand=True, pady=(0, 20))

        lbl_hero_title = tk.Label(
            header_card,
            text="ℹ️ ABOUT INTERACTIVE PDF FORM SUITE",
            font=("Segoe UI", 13, "bold"),
            fg="#FFFFFF", bg="#1F4E78"
        )
        lbl_hero_title.pack(anchor="w")

        lbl_hero_desc = tk.Label(
            header_card,
            text="Standalone, Schema-Driven Interactive Fillable PDF Form Generator & Batch Data Extraction Pipeline",
            font=("Segoe UI", 9, "italic"),
            fg="#DDEBF7", bg="#1F4E78"
        )
        lbl_hero_desc.pack(anchor="w", pady=(4, 0))

        # 2. KEY USES & CORE CAPABILITIES SECTION
        lbl_uses_header = tk.Label(
            scroll_frame,
            text="CORE FEATURES:",
            font=("Segoe UI", 11, "bold"),
            fg="#1F4E78", bg="#F4F6F9"
        )
        lbl_uses_header.pack(anchor="w", pady=(0, 10))

        capabilities = [
            (
                "📋 Schema-Driven Form Builder",
                "Create interactive PDF forms from Excel or AI JSON."
            ),
            (
                "🎨 Automatic Professional Layouts",
                "Multi-column layouts, themes, validation, and field controls."
            ),
            (
                "📝 Built-in Form Utilities",
                "Generate form summaries, validate entries, reset forms, and lock completed forms."
            ),
            (
                "📊 Batch PDF Consolidation",
                "Extract data from completed PDFs into Excel reports."
            ),
            (
                "🤖 AI-Ready Configuration",
                "Generate complete form templates using modern AI assistants."
            ),
            (
                "🔒 Offline & Portable",
                "No cloud, no database, secure standalone application."
            )
        ]

        for title_text, body_text in capabilities:
            card = tk.Frame(scroll_frame, bg="#FFFFFF", padx=15, pady=12, relief="solid", bd=1)
            card.pack(fill="x", expand=True, pady=6)

            lbl_c_title = tk.Label(
                card,
                text=title_text,
                font=("Segoe UI", 10, "bold"),
                fg="#1F4E78", bg="#FFFFFF"
            )
            lbl_c_title.pack(anchor="w")

            lbl_c_body = tk.Label(
                card,
                text=body_text,
                font=("Segoe UI", 9),
                fg="#333333", bg="#FFFFFF",
                wraplength=750, justify="left"
            )
            lbl_c_body.pack(anchor="w", pady=(3, 0))

        # 3. QUICK ACTION BUTTONS FOOTER
        footer_frame = tk.Frame(scroll_frame, bg="#F4F6F9", pady=15)
        footer_frame.pack(fill="x", expand=True)

        btn_guide = tk.Button(
            footer_frame,
            text="📘 Open User Guide",
            font=("Segoe UI", 9, "bold"),
            bg="#DDEBF7", fg="#1F4E78",
            relief="solid", bd=1, padx=12, pady=6, cursor="hand2",
            command=self.open_user_guide_in_browser
        )
        btn_guide.pack(side="left", padx=(0, 10))

        btn_create_demo = tk.Button(
            footer_frame,
            text="⚡ Compile Sample Form",
            font=("Segoe UI", 9, "bold"),
            bg="#1F4E78", fg="#FFFFFF",
            relief="solid", bd=1, padx=12, pady=6, cursor="hand2",
            command=self.compile_fillable_pdf
        )
        btn_create_demo.pack(side="left", padx=5)

    # =========================================================================
    # TAB 1: Global Basic Form Details (Sheet 1 Editor)
    # =========================================================================
    def _build_tab1_basic_details(self, parent):
        box = tk.LabelFrame(parent, text=" Global Form Header Banners & Theme Settings ", font=("Segoe UI", 11, "bold"), bg="#F4F6F9", fg="#1F4E78", padx=15, pady=12)
        box.pack(fill="both", expand=True, padx=10, pady=10)

        # Form Title
        r1 = tk.Frame(box, bg="#F4F6F9")
        r1.pack(fill="x", pady=5)
        tk.Label(r1, text="Form Title (Header Banner Line 1):", font=("Segoe UI", 10, "bold"), bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_title = tk.Entry(r1, font=("Segoe UI", 10), bg="#FFFFFF", relief="solid", bd=1)
        self.ent_title.insert(0, self.form_settings["Form_Title"])
        self.ent_title.pack(side="left", fill="x", expand=True)

        # Form Subtitle
        r2 = tk.Frame(box, bg="#F4F6F9")
        r2.pack(fill="x", pady=5)
        tk.Label(r2, text="Form Subtitle (Header Banner Line 2):", font=("Segoe UI", 10, "bold"), bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_subtitle = tk.Entry(r2, font=("Segoe UI", 10), bg="#FFFFFF", relief="solid", bd=1)
        self.ent_subtitle.insert(0, self.form_settings["Form_Subtitle"])
        self.ent_subtitle.pack(side="left", fill="x", expand=True)

        # Instruction Point 1 (Mandatory Locked System Rule)
        r3 = tk.Frame(box, bg="#F4F6F9")
        r3.pack(fill="x", pady=5)
        tk.Label(r3, text="Instruction Point 1 [Locked System Rule]:", font=("Segoe UI", 10, "bold"), fg="#C00000", bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_inst1 = tk.Entry(r3, font=("Segoe UI", 9, "bold"), bg="#E2E2E2", fg="#C00000", disabledbackground="#E2E2E2", disabledforeground="#C00000", relief="solid", bd=1)
        self.ent_inst1.insert(0, self.form_settings["Instruction_Point_1"])
        self.ent_inst1.configure(state="disabled")
        self.ent_inst1.pack(side="left", fill="x", expand=True)

        # Instruction Point 2 (Mandatory Locked System Rule)
        r4 = tk.Frame(box, bg="#F4F6F9")
        r4.pack(fill="x", pady=5)
        tk.Label(r4, text="Instruction Point 2 [Locked System Rule]:", font=("Segoe UI", 10, "bold"), fg="#C00000", bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_inst2 = tk.Entry(r4, font=("Segoe UI", 9, "bold"), bg="#E2E2E2", fg="#C00000", disabledbackground="#E2E2E2", disabledforeground="#C00000", relief="solid", bd=1)
        self.ent_inst2.insert(0, self.form_settings["Instruction_Point_2"])
        self.ent_inst2.configure(state="disabled")
        self.ent_inst2.pack(side="left", fill="x", expand=True)

        # Instruction Point 3 (User Configurable)
        r5 = tk.Frame(box, bg="#F4F6F9")
        r5.pack(fill="x", pady=5)
        tk.Label(r5, text="Instruction Point 3:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_inst3 = tk.Entry(r5, font=("Segoe UI", 10), bg="#FFFFFF", relief="solid", bd=1)
        self.ent_inst3.insert(0, self.form_settings.get("Instruction_Point_3", ""))
        self.ent_inst3.pack(side="left", fill="x", expand=True)

        # Instruction Point 4 (User Configurable Optional)
        r5_4 = tk.Frame(box, bg="#F4F6F9")
        r5_4.pack(fill="x", pady=5)
        tk.Label(r5_4, text="Instruction Point 4 (Optional):", font=("Segoe UI", 10, "bold"), bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_inst4 = tk.Entry(r5_4, font=("Segoe UI", 10), bg="#FFFFFF", relief="solid", bd=1)
        self.ent_inst4.insert(0, self.form_settings.get("Instruction_Point_4", ""))
        self.ent_inst4.pack(side="left", fill="x", expand=True)

        # Footer Text
        r6 = tk.Frame(box, bg="#F4F6F9")
        r6.pack(fill="x", pady=5)
        tk.Label(r6, text="Page Footer Banner Text:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_footer = tk.Entry(r6, font=("Segoe UI", 10), bg="#FFFFFF", relief="solid", bd=1)
        self.ent_footer.insert(0, self.form_settings["Footer_Text"])
        self.ent_footer.pack(side="left", fill="x", expand=True)

        # Initial Lock Status
        r7 = tk.Frame(box, bg="#F4F6F9")
        r7.pack(fill="x", pady=5)
        tk.Label(r7, text="Default Form Lock Status:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.cmb_lock = ttk.Combobox(r7, values=["Locked", "Unlocked", "Not Required"], font=("Segoe UI", 10), width=20, state="readonly")
        self.cmb_lock.set(self.form_settings["Default_Lock_Status"])
        self.cmb_lock.pack(side="left")

        # Summary Section Title (Locked System Default)
        r8 = tk.Frame(box, bg="#F4F6F9")
        r8.pack(fill="x", pady=5)
        tk.Label(r8, text="Summary Section Title [Locked]:", font=("Segoe UI", 10, "bold"), fg="#1F4E78", bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_sum_sec_title = tk.Entry(r8, font=("Segoe UI", 9, "bold"), bg="#E2E2E2", fg="#1F4E78", disabledbackground="#E2E2E2", disabledforeground="#1F4E78", relief="solid", bd=1)
        self.ent_sum_sec_title.insert(0, "GENERATE FORM SUMMARY & COPY TEXT")
        self.ent_sum_sec_title.configure(state="disabled")
        self.ent_sum_sec_title.pack(side="left", fill="x", expand=True)

        # Summary Button Text (Locked System Default)
        r9 = tk.Frame(box, bg="#F4F6F9")
        r9.pack(fill="x", pady=5)
        tk.Label(r9, text="Summary Action Button Text [Locked]:", font=("Segoe UI", 10, "bold"), fg="#1F4E78", bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_sum_btn_txt = tk.Entry(r9, font=("Segoe UI", 9, "bold"), bg="#E2E2E2", fg="#1F4E78", disabledbackground="#E2E2E2", disabledforeground="#1F4E78", relief="solid", bd=1)
        self.ent_sum_btn_txt.insert(0, ">>>> Click the Button - To Generate Form Summary <<<<")
        self.ent_sum_btn_txt.configure(state="disabled")
        self.ent_sum_btn_txt.pack(side="left", fill="x", expand=True)

        # Summary Field Label
        r10 = tk.Frame(box, bg="#F4F6F9")
        r10.pack(fill="x", pady=5)
        tk.Label(r10, text="Summary Output Field Label:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9", width=32, anchor="w").pack(side="left")
        self.ent_sum_fld_lbl = tk.Entry(r10, font=("Segoe UI", 10), bg="#FFFFFF", relief="solid", bd=1)
        self.ent_sum_fld_lbl.insert(0, self.form_settings.get("Summary_Field_Label", ""))
        self.ent_sum_fld_lbl.pack(side="left", fill="x", expand=True)

    # =========================================================================
    # TAB 2: Section Details & Grid Layout Columns (Sheet 2 Editor)
    # =========================================================================
    def _build_tab2_sections(self, parent):
        box = tk.LabelFrame(parent, text=" Section Layout & Grid Column Architecture ", font=("Segoe UI", 11, "bold"), bg="#F4F6F9", fg="#1F4E78", padx=12, pady=10)
        box.pack(fill="both", expand=True, padx=10, pady=10)

        # Section Grid Table
        cols = ("Section_Number", "Section_ID", "Section_Heading_Title", "Layout_Grid_Columns")
        self.sec_tree = ttk.Treeview(box, columns=cols, show="headings", height=7)

        self.sec_tree.heading("Section_Number", text="Sec # (Locked)")
        self.sec_tree.column("Section_Number", width=110, anchor="center")

        self.sec_tree.heading("Section_ID", text="Section_ID (Protected)")
        self.sec_tree.column("Section_ID", width=140, anchor="center")

        self.sec_tree.heading("Section_Heading_Title", text="Section Heading Title (Editable)")
        self.sec_tree.column("Section_Heading_Title", width=420, anchor="w")

        self.sec_tree.heading("Layout_Grid_Columns", text="Layout Grid Columns (Select 1-4)")
        self.sec_tree.column("Layout_Grid_Columns", width=200, anchor="center")

        sec_scroll = ttk.Scrollbar(box, orient="vertical", command=self.sec_tree.yview)
        self.sec_tree.configure(yscrollcommand=sec_scroll.set)

        sec_scroll.pack(side="right", fill="y")
        self.sec_tree.pack(fill="both", expand=True, side="top", pady=(0, 10))
        self.sec_tree.bind("<<TreeviewSelect>>", self.on_section_selected)

        # Section Edit Inputs (LOCKED Section_ID!)
        edit_box = tk.Frame(box, bg="#F4F6F9")
        edit_box.pack(fill="x", side="bottom")

        r1 = tk.Frame(edit_box, bg="#F4F6F9")
        r1.pack(fill="x", pady=4)

        tk.Label(r1, text="Section ID (Protected):", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.ent_sec_id = tk.Entry(r1, font=("Segoe UI", 10), width=12, bg="#F8FAFC", relief="solid", bd=1, state="readonly")
        self.ent_sec_id.pack(side="left", padx=(5, 15))

        tk.Label(r1, text="Section Heading Title:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.ent_sec_title = tk.Entry(r1, font=("Segoe UI", 10), width=48, bg="#FFFFFF", relief="solid", bd=1)
        self.ent_sec_title.pack(side="left", padx=(5, 15))

        tk.Label(r1, text="Grid Columns:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.cmb_sec_cols = ttk.Combobox(r1, values=["1", "2", "3", "4"], font=("Segoe UI", 10), width=6, state="readonly")
        self.cmb_sec_cols.set("3")
        self.cmb_sec_cols.pack(side="left", padx=5)

        r2 = tk.Frame(edit_box, bg="#F4F6F9")
        r2.pack(fill="x", pady=8)

        btn_add_sec = tk.Button(r2, text="➕ Add Section", font=("Segoe UI", 10, "bold"), bg="#DDEBF7", fg="#1F4E78", relief="solid", bd=1, padx=12, pady=3, command=self.add_section_row)
        btn_add_sec.pack(side="left", padx=(0, 10))

        btn_upd_sec = tk.Button(r2, text="✏️ Update Selected Section", font=("Segoe UI", 10, "bold"), bg="#E2E2E2", fg="#333333", relief="solid", bd=1, padx=12, pady=3, command=self.update_section_row)
        btn_upd_sec.pack(side="left", padx=10)

        btn_del_sec = tk.Button(r2, text="❌ Delete Selected Section", font=("Segoe UI", 10, "bold"), bg="#FFC7CE", fg="#C00000", relief="solid", bd=1, padx=12, pady=3, command=self.delete_section_row)
        btn_del_sec.pack(side="left", padx=10)

        self.refresh_sections_treeview()

    def refresh_sections_treeview(self, select_idx=None):
        """Refreshes Section Treeview grid and maintains selection focus if provided."""
        for item in self.sec_tree.get_children():
            self.sec_tree.delete(item)

        for row in self.section_rows:
            self.sec_tree.insert("", "end", values=(row[0], row[1], row[2], row[3]))

        # Maintain Row Selection Focus after Update
        children = self.sec_tree.get_children()
        if select_idx is not None and 0 <= select_idx < len(children):
            target_item = children[select_idx]
            self.sec_tree.selection_set(target_item)
            self.sec_tree.focus(target_item)
            self.sec_tree.see(target_item)

        # Refresh Section ID Dropdown in Tab 3
        sec_ids = [row[1] for row in self.section_rows]
        if hasattr(self, 'cmb_fld_sec'):
            self.cmb_fld_sec.configure(values=sec_ids)

    def on_section_selected(self, event):
        """Populates edit inputs when section row is selected."""
        sel = self.sec_tree.selection()
        if sel:
            vals = self.sec_tree.item(sel[0], "values")
            self.ent_sec_id.configure(state="normal")
            self.ent_sec_id.delete(0, tk.END)
            self.ent_sec_id.insert(0, vals[1])
            self.ent_sec_id.configure(state="readonly")
            
            self.ent_sec_title.delete(0, tk.END)
            self.ent_sec_title.insert(0, vals[2])
            self.cmb_sec_cols.set(vals[3])

    def add_section_row(self):
        num = len(self.section_rows) + 1
        sec_id = f"SEC_{num}"
        sec_title = self.ent_sec_title.get().strip()
        sec_cols = self.cmb_sec_cols.get().strip()

        if not sec_title:
            self.set_inline_status("Error: Section Heading Title cannot be empty!", is_error=True)
            return

        self.section_rows.append([num, sec_id, sec_title, sec_cols])
        new_idx = len(self.section_rows) - 1
        self.refresh_sections_treeview(select_idx=new_idx)
        self.set_inline_status(f"Added new Section: '{sec_id}' with {sec_cols} grid columns.")

    def update_section_row(self):
        sel = self.sec_tree.selection()
        if not sel:
            self.set_inline_status("Select a section row to update.", is_error=True)
            return

        idx = self.sec_tree.index(sel[0])
        sec_id = self.section_rows[idx][1]
        sec_title = self.ent_sec_title.get().strip()
        sec_cols = self.cmb_sec_cols.get().strip()

        self.section_rows[idx] = [idx + 1, sec_id, sec_title, sec_cols]
        # Preserve active row selection!
        self.refresh_sections_treeview(select_idx=idx)
        self.set_inline_status(f"Updated Section: '{sec_id}'.")

    def delete_section_row(self):
        sel = self.sec_tree.selection()
        if not sel:
            self.set_inline_status("Select a section row to delete.", is_error=True)
            return

        idx = self.sec_tree.index(sel[0])
        del self.section_rows[idx]
        # Re-number remaining sections and systematically re-id
        for i in range(len(self.section_rows)):
            self.section_rows[i][0] = i + 1
            self.section_rows[i][1] = f"SEC_{i + 1}"
        self.refresh_sections_treeview()
        self.set_inline_status("Deleted selected section row.")

    # =========================================================================
    # TAB 3: Field Configuration Schema (Sheet 3 Editor)
    # =========================================================================
    def _build_tab3_fields(self, parent):
        box = tk.LabelFrame(parent, text=" Field Configuration Schema Data Grid ", font=("Segoe UI", 11, "bold"), bg="#F4F6F9", fg="#1F4E78", padx=12, pady=10)
        box.pack(fill="both", expand=True, padx=10, pady=10)

        cols = ("Field_ID", "Section_ID", "Field_Label", "Field_DataType", "Default_Value", "Validation_Operator", "Validation_Param", "Dropdown_Options")
        self.fld_tree = ttk.Treeview(box, columns=cols, show="headings", height=8)

        col_widths = {
            "Field_ID": 100, "Section_ID": 90, "Field_Label": 170, "Field_DataType": 100,
            "Default_Value": 110, "Validation_Operator": 140, "Validation_Param": 140, "Dropdown_Options": 210
        }
        for col in cols:
            text_hdr = col
            if col in ("Field_ID", "Section_ID"):
                text_hdr = f"{col} (Protected)"
            elif col in ("Validation_Operator", "Validation_Param", "Dropdown_Options"):
                text_hdr = f"{col} (Conditional)"
            self.fld_tree.heading(col, text=text_hdr)
            self.fld_tree.column(col, width=col_widths.get(col, 110), anchor="w")

        fld_scroll = ttk.Scrollbar(box, orient="vertical", command=self.fld_tree.yview)
        self.fld_tree.configure(yscrollcommand=fld_scroll.set)

        fld_scroll.pack(side="right", fill="y")
        self.fld_tree.pack(fill="both", expand=True, side="top", pady=(0, 10))
        self.fld_tree.bind("<<TreeviewSelect>>", self.on_field_selected)

        # Field Inputs Frame (PROTECTED Field_ID and Section_ID!)
        f_edit = tk.Frame(box, bg="#F4F6F9")
        f_edit.pack(fill="x", side="bottom")

        r1 = tk.Frame(f_edit, bg="#F4F6F9")
        r1.pack(fill="x", pady=3)

        tk.Label(r1, text="Field ID (Protected):", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.ent_fld_id = tk.Entry(r1, font=("Segoe UI", 10), width=12, bg="#F8FAFC", relief="solid", bd=1, state="readonly")
        self.ent_fld_id.pack(side="left", padx=(2, 12))

        tk.Label(r1, text="Parent Sec ID (Protected):", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        sec_ids = [row[1] for row in self.section_rows]
        self.cmb_fld_sec = ttk.Combobox(r1, values=sec_ids, font=("Segoe UI", 10), width=10, state="readonly")
        if sec_ids: self.cmb_fld_sec.set(sec_ids[0])
        self.cmb_fld_sec.pack(side="left", padx=(2, 12))

        tk.Label(r1, text="Label Heading:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.ent_fld_lbl = tk.Entry(r1, font=("Segoe UI", 10), width=28, bg="#FFFFFF", relief="solid", bd=1)
        self.ent_fld_lbl.pack(side="left", padx=(2, 12))

        tk.Label(r1, text="DataType:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.cmb_fld_dtype = ttk.Combobox(r1, values=["short_text", "long_text", "whole_number", "decimal", "choice", "multi_choice", "radio_choice", "date"], font=("Segoe UI", 10), width=12, state="readonly")
        self.cmb_fld_dtype.set("short_text")
        self.cmb_fld_dtype.pack(side="left", padx=2)
        self.cmb_fld_dtype.bind("<<ComboboxSelected>>", self.on_datatype_changed)

        r2 = tk.Frame(f_edit, bg="#F4F6F9")
        r2.pack(fill="x", pady=3)

        tk.Label(r2, text="Default Val:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.ent_fld_def = tk.Entry(r2, font=("Segoe UI", 10), width=14, bg="#FFFFFF", relief="solid", bd=1)
        self.ent_fld_def.pack(side="left", padx=(2, 10))

        tk.Label(r2, text="Operator (Conditional):", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.cmb_fld_op = ttk.Combobox(r2, values=self.operator_matrix["short_text"], font=("Segoe UI", 10), width=12, state="readonly")
        self.cmb_fld_op.set("none")
        self.cmb_fld_op.pack(side="left", padx=(2, 10))
        self.cmb_fld_op.bind("<<ComboboxSelected>>", self.on_operator_changed)

        tk.Label(r2, text="Param Val (Conditional):", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.ent_fld_param = tk.Entry(r2, font=("Segoe UI", 10), width=12, bg="#F8FAFC", relief="solid", bd=1, state="disabled")
        self.ent_fld_param.pack(side="left", padx=(2, 8))

        lbl_param_hint = tk.Label(
            r2,
            text="*(Param Val required only for operators min_length, max_length, comparison, range)",
            font=("Segoe UI", 8, "italic"),
            fg="#595959",
            bg="#F4F6F9"
        )
        lbl_param_hint.pack(side="left")

        # Choice Options Entry + Clear Note
        r_opts = tk.Frame(f_edit, bg="#F4F6F9")
        r_opts.pack(fill="x", pady=3)

        tk.Label(r_opts, text="Choice Options:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(side="left")
        self.ent_fld_opts = tk.Entry(r_opts, font=("Segoe UI", 10), width=45, bg="#F8FAFC", relief="solid", bd=1, state="disabled")
        self.ent_fld_opts.pack(side="left", padx=(5, 10))

        lbl_note = tk.Label(
            r_opts,
            text="*(Note: Enter comma separated values e.g. Option 1, Option 2, Option 3)",
            font=("Segoe UI", 9, "italic"),
            fg="#1F4E78",
            bg="#F4F6F9"
        )
        lbl_note.pack(side="left")

        r3 = tk.Frame(f_edit, bg="#F4F6F9")
        r3.pack(fill="x", pady=8)

        btn_add_fld = tk.Button(r3, text="➕ Add Field", font=("Segoe UI", 10, "bold"), bg="#DDEBF7", fg="#1F4E78", relief="solid", bd=1, padx=12, pady=3, command=self.add_field_row)
        btn_add_fld.pack(side="left", padx=(0, 10))

        btn_upd_fld = tk.Button(r3, text="✏️ Update Selected Field", font=("Segoe UI", 10, "bold"), bg="#E2E2E2", fg="#333333", relief="solid", bd=1, padx=12, pady=3, command=self.update_field_row)
        btn_upd_fld.pack(side="left", padx=10)

        btn_del_fld = tk.Button(r3, text="❌ Delete Selected Field", font=("Segoe UI", 10, "bold"), bg="#FFC7CE", fg="#C00000", relief="solid", bd=1, padx=12, pady=3, command=self.delete_field_row)
        btn_del_fld.pack(side="left", padx=10)

        self.refresh_fields_treeview()

    def refresh_fields_treeview(self, select_idx=None):
        """Refreshes Field Treeview grid and maintains selection focus if provided."""
        for item in self.fld_tree.get_children():
            self.fld_tree.delete(item)

        for row in self.field_schema_rows:
            display_row = (row[0], row[1], row[2], row[3], row[4], row[6], row[7], row[8])
            self.fld_tree.insert("", "end", values=display_row)

        # Maintain Row Selection Focus after Update!
        children = self.fld_tree.get_children()
        if select_idx is not None and 0 <= select_idx < len(children):
            target_item = children[select_idx]
            self.fld_tree.selection_set(target_item)
            self.fld_tree.focus(target_item)
            self.fld_tree.see(target_item)

    def on_datatype_changed(self, event=None):
        dtype = self.cmb_fld_dtype.get().strip().lower()
        allowed_ops = self.operator_matrix.get(dtype, ["none"])
        self.cmb_fld_op.configure(values=allowed_ops)
        if self.cmb_fld_op.get() not in allowed_ops:
            self.cmb_fld_op.set(allowed_ops[0])

        # CONTEXTUAL LOCKING: Enable Choice Options for choice, multi_choice, or radio_choice datatypes!
        if dtype in ("choice", "multi_choice", "radio_choice"):
            self.ent_fld_opts.configure(state="normal", bg="#FFFFFF")
        else:
            self.ent_fld_opts.delete(0, tk.END)
            self.ent_fld_opts.configure(state="disabled", bg="#F8FAFC")

        self.on_operator_changed()

    def on_operator_changed(self, event=None):
        op = self.cmb_fld_op.get().strip().lower()
        # CONTEXTUAL LOCKING: Enable Param Val ONLY when operator requires parameter!
        requires_param = op not in ("none", "date_ddmmyyyy", "")
        if requires_param:
            self.ent_fld_param.configure(state="normal", bg="#FFFFFF")
        else:
            self.ent_fld_param.delete(0, tk.END)
            self.ent_fld_param.configure(state="disabled", bg="#F8FAFC")

    def on_field_selected(self, event):
        sel = self.fld_tree.selection()
        if sel:
            vals = self.fld_tree.item(sel[0], "values")
            self.ent_fld_id.configure(state="normal")
            self.ent_fld_id.delete(0, tk.END); self.ent_fld_id.insert(0, vals[0])
            self.ent_fld_id.configure(state="readonly")

            self.cmb_fld_sec.set(vals[1])
            self.ent_fld_lbl.delete(0, tk.END); self.ent_fld_lbl.insert(0, vals[2])
            self.cmb_fld_dtype.set(vals[3])
            self.on_datatype_changed()

            self.ent_fld_def.delete(0, tk.END); self.ent_fld_def.insert(0, vals[4])
            self.cmb_fld_op.set(vals[5])
            self.on_operator_changed()

            if vals[6]:
                self.ent_fld_param.configure(state="normal")
                self.ent_fld_param.delete(0, tk.END); self.ent_fld_param.insert(0, vals[6])

            if vals[7]:
                self.ent_fld_opts.configure(state="normal")
                self.ent_fld_opts.delete(0, tk.END); self.ent_fld_opts.insert(0, vals[7])

    def validate_and_clean_choice_fields(self, dtype: str, def_val: str, opts: str):
        """
        Validates choice options & default values to prevent ReportLab choice errors!
        Returns (clean_def_val, clean_opts, warning_msg).
        """
        if dtype != "choice":
            return def_val, opts, None

        opt_list = [o.strip() for o in opts.split(",") if o.strip()]
        if not opt_list:
            opt_list = ["Option 1", "Option 2"]
            opts = "Option 1, Option 2"

        warning = None
        if def_val and def_val not in opt_list:
            warning = f"Notice: Default Value ('{def_val}') was not in Choice Options ({opt_list}). Automatically reset Default Value to blank."
            def_val = ""

        return def_val, opts, warning

    def add_field_row(self):
        sec_id = self.cmb_fld_sec.get().strip()
        label = self.ent_fld_lbl.get().strip()
        dtype = self.cmb_fld_dtype.get().strip().lower()
        def_val = self.ent_fld_def.get().strip()
        op = self.cmb_fld_op.get().strip().lower()
        param = self.ent_fld_param.get().strip()
        opts = self.ent_fld_opts.get().strip()

        if not label:
            self.set_inline_status("Error: Label Heading cannot be empty!", is_error=True)
            return

        def_val, opts, warn_msg = self.validate_and_clean_choice_fields(dtype, def_val, opts)

        # Systematically generate Field_ID based on section
        sec_num = sec_id.replace("SEC_", "")
        sub_count = sum(1 for r in self.field_schema_rows if r[1] == sec_id) + 1
        f_id = f"Field_{sec_num}_{sub_count}"

        row = [f_id, sec_id, label, dtype, def_val, "", op, param, opts]
        self.field_schema_rows.append(row)
        new_idx = len(self.field_schema_rows) - 1
        self.refresh_fields_treeview(select_idx=new_idx)

        status_txt = f"Added new Field: '{f_id}' ({dtype})."
        if warn_msg:
            status_txt += f" {warn_msg}"
        self.set_inline_status(status_txt)

    def update_field_row(self):
        sel = self.fld_tree.selection()
        if not sel:
            self.set_inline_status("Select a field row to update.", is_error=True)
            return

        idx = self.fld_tree.index(sel[0])
        f_id = self.field_schema_rows[idx][0]
        sec_id = self.cmb_fld_sec.get().strip()
        label = self.ent_fld_lbl.get().strip()
        dtype = self.cmb_fld_dtype.get().strip().lower()
        def_val = self.ent_fld_def.get().strip()
        op = self.cmb_fld_op.get().strip().lower()
        param = self.ent_fld_param.get().strip()
        opts = self.ent_fld_opts.get().strip()

        def_val, opts, warn_msg = self.validate_and_clean_choice_fields(dtype, def_val, opts)

        self.field_schema_rows[idx] = [f_id, sec_id, label, dtype, def_val, "", op, param, opts]
        # Preserve active row selection after update!
        self.refresh_fields_treeview(select_idx=idx)

        status_txt = f"Updated Field: '{f_id}'."
        if warn_msg:
            status_txt += f" {warn_msg}"
        self.set_inline_status(status_txt)

    def delete_field_row(self):
        sel = self.fld_tree.selection()
        if not sel:
            self.set_inline_status("Select a field row to delete.", is_error=True)
            return

        idx = self.fld_tree.index(sel[0])
        del self.field_schema_rows[idx]
        self.refresh_fields_treeview()
        self.set_inline_status("Deleted selected field row.")

    # =========================================================================
    # TAB 4: Batch PDF Data Consolidator (Medium Blue Button!)
    # =========================================================================
    def _build_tab4_consolidator(self, parent):
        box = tk.LabelFrame(parent, text=" Batch PDF Data Extraction & Excel Report Export ", font=("Segoe UI", 11, "bold"), bg="#F4F6F9", fg="#1F4E78", padx=15, pady=15)
        box.pack(fill="x", padx=10, pady=15)

        tk.Label(box, text="Target PDF Directory containing completed forms:", font=("Segoe UI", 10, "bold"), bg="#F4F6F9").pack(anchor="w", pady=(0, 5))

        r_path = tk.Frame(box, bg="#F4F6F9")
        r_path.pack(fill="x", pady=5)

        self.ent_pdf_dir = tk.Entry(r_path, font=("Segoe UI", 10), bg="#FFFFFF", relief="solid", bd=1)
        self.ent_pdf_dir.insert(0, self.output_pdf_folder)
        self.ent_pdf_dir.pack(side="left", fill="x", expand=True, padx=(0, 10))

        btn_browse = tk.Button(
            r_path, text="📁 Browse Folder", font=("Segoe UI", 10, "bold"),
            bg="#E2E2E2", fg="#333333", relief="solid", bd=1, padx=12, pady=2,
            command=self.browse_pdf_folder
        )
        btn_browse.pack(side="right")

        # Medium-sized Navy Blue Button (Left aligned, fixed padding instead of full-width fill="x")
        btn_consolidate = tk.Button(
            box, text="📊 Run Data Consolidation into Excel Report",
            font=("Segoe UI", 10, "bold"), bg="#1F4E78", fg="#FFFFFF",
            activebackground="#143451", activeforeground="#FFFFFF",
            relief="solid", bd=1, padx=25, pady=8, cursor="hand2",
            command=self.run_data_consolidation
        )
        btn_consolidate.pack(anchor="w", pady=(15, 8))

        self.lbl_consol_result = tk.Label(
            box, text="No consolidation run yet. Click above button to process PDFs.",
            font=("Segoe UI", 10, "italic"), bg="#F4F6F9", fg="#595959", anchor="w"
        )
        self.lbl_consol_result.pack(fill="x", pady=(5, 0))

    # =========================================================================
    # TAB 5: Activity Logs & Diagnostics
    # =========================================================================
    def _build_tab5_logs(self, parent):
        box = tk.LabelFrame(parent, text=" Live Console Output & Session Diagnostics ", font=("Segoe UI", 11, "bold"), bg="#F4F6F9", fg="#1F4E78", padx=10, pady=10)
        box.pack(fill="both", expand=True, padx=10, pady=10)

        self.txt_console = tk.Text(
            box, wrap="word", font=("Consolas", 10),
            bg="#1E1E1E", fg="#D4D4D4", insertbackground="white", relief="flat"
        )
        scrollbar = ttk.Scrollbar(box, orient="vertical", command=self.txt_console.yview)
        self.txt_console.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        self.txt_console.pack(side="left", fill="both", expand=True)

        toolbar = tk.Frame(parent, bg="#F4F6F9", pady=6)
        toolbar.pack(fill="x", padx=10)

        btn_view_file = tk.Button(
            toolbar, text="📄 Open Session Log File in Text Editor", font=("Segoe UI", 10, "bold"),
            bg="#E2E2E2", fg="#333333", relief="solid", bd=1, padx=12, pady=3,
            command=self.open_log_file
        )
        btn_view_file.pack(side="left")

        btn_open_guide = tk.Button(
            toolbar, text="🌐 Open HTML5 User Guide in Browser", font=("Segoe UI", 10, "bold"),
            bg="#1F4E78", fg="#FFFFFF", activebackground="#143451", activeforeground="#FFFFFF",
            relief="solid", bd=1, padx=15, pady=3, cursor="hand2",
            command=self.open_user_guide_in_browser
        )
        btn_open_guide.pack(side="left", padx=10)

    def open_user_guide_in_browser(self):
        """Opens HELP_USER_GUIDE.html in the default web browser."""
        base_dir = get_base_dir()
        guide_path = os.path.join(base_dir, "HELP_USER_GUIDE.html")
        if not os.path.exists(guide_path):
            guide_path = os.path.join(os.path.dirname(base_dir), "HELP_USER_GUIDE.html")

        if os.path.exists(guide_path):
            abs_guide = os.path.abspath(guide_path)
            webbrowser.open(abs_guide)
            self.set_inline_status(f"Opening User Guide in default browser: '{abs_guide}'", is_success=True)
            self.log_info(f"User Guide opened in web browser: '{abs_guide}'")
        else:
            self.set_inline_status("User Guide file 'HELP_USER_GUIDE.html' not found on disk.", is_error=True)

    # =========================================================================
    # CORE PDF COMPILER & STATE SYNC LOGIC
    # =========================================================================
    def sync_ui_to_settings(self):
        """Reads current values from UI entries into internal data model."""
        self.form_settings["Form_Title"] = self.ent_title.get().strip()
        self.form_settings["Form_Subtitle"] = self.ent_subtitle.get().strip()
        self.form_settings["Instruction_Point_1"] = "!! Save As - this PDF file before mailing. Cross-check whether all filled form data is properly saved before sending. !!"
        self.form_settings["Instruction_Point_2"] = "!! Before saving the file Click 'Generate Summary' Button to update it..!!   !! Use Adobe Acrobat Reader software for filling this form. !!"
        self.form_settings["Instruction_Point_3"] = self.ent_inst3.get().strip() if hasattr(self, 'ent_inst3') else ""
        self.form_settings["Instruction_Point_4"] = self.ent_inst4.get().strip() if hasattr(self, 'ent_inst4') else ""
        self.form_settings["Footer_Text"] = self.ent_footer.get().strip()
        self.form_settings["Default_Lock_Status"] = self.cmb_lock.get().strip()
        if hasattr(self, 'ent_sum_sec_title'):
            self.form_settings["Summary_Section_Title"] = "GENERATE FORM SUMMARY & COPY TEXT"
            self.form_settings["Summary_Button_Text"] = ">>>> Click the Button - To Generate Form Summary <<<<"
            self.form_settings["Summary_Field_Label"] = self.ent_sum_fld_lbl.get().strip() if hasattr(self, 'ent_sum_fld_lbl') else "GENERATED FORM SUMMARY TEXT"

    def compile_fillable_pdf(self):
        """
        Validates native Tkinter UI state, syncs configuration, and compiles the PDF form!
        """
        self.sync_ui_to_settings()
        self.set_inline_status("Validating form configuration...", is_error=False)

        # 1. Validation Checks
        if not self.form_settings["Form_Title"]:
            self.set_inline_status("Validation Failed: Form Title cannot be empty!", is_error=True)
            self.log_error("Validation Failed: Form Title is empty.")
            return

        if len(self.form_settings["Form_Title"]) > 60:
            self.set_inline_status(f"Validation Failed: Form Title length ({len(self.form_settings['Form_Title'])} chars) exceeds max limit of 60.", is_error=True)
            return

        if len(self.form_settings.get("Form_Subtitle", "")) > 85:
            self.set_inline_status(f"Validation Failed: Form Subtitle length ({len(self.form_settings['Form_Subtitle'])} chars) exceeds max limit of 85.", is_error=True)
            return

        for k, name in [("Instruction_Point_1", "Instruction Point 1"), ("Instruction_Point_2", "Instruction Point 2"), ("Instruction_Point_3", "Instruction Point 3"), ("Footer_Text", "Footer Text"), ("Summary_Section_Title", "Summary Section Title"), ("Summary_Button_Text", "Summary Button Text"), ("Summary_Field_Label", "Summary Field Label")]:
            val = self.form_settings.get(k, "")
            max_l = 180 if k in ("Instruction_Point_1", "Instruction_Point_2") else 110
            if len(val) > max_l:
                self.set_inline_status(f"Validation Failed: {name} length ({len(val)} chars) exceeds max limit of {max_l}.", is_error=True)
                return

        # 2. Transparently write current UI state to master Excel configurator file
        try:
            self.save_all_sheets_to_excel()
        except PermissionError:
            self.set_inline_status("Permission Error: 'fillable_pdf_configurator.xlsx' is open in Excel! Please close Excel and try again.", is_error=True)
            return
        except Exception as e:
            self.set_inline_status(f"Error saving configuration: {e}", is_error=True)
            return

        # 3. Generate Interactive PDF Form
        self.log_info("Compiling Fillable PDF Form...")
        try:
            pdf_path = generate_pdf_from_excel(self.excel_path)
            abs_pdf = os.path.abspath(pdf_path)
            msg = f"SUCCESS! Fillable PDF Form Created at: '{abs_pdf}'"
            self.set_inline_status(msg, is_success=True)
            self.log_info(msg)

            # Auto-open generated PDF in default viewer
            try:
                os.startfile(abs_pdf)
            except Exception:
                pass

        except Exception as e:
            err_msg = f"PDF Generation Error: {e}"
            self.set_inline_status(err_msg, is_error=True)
            self.log_error(f"{err_msg}\n{traceback.format_exc()}")

    def export_excel_dialog(self):
        """Prompt user for destination file and export full 3-sheet Excel configuration."""
        self.sync_ui_to_settings()
        dest_file = filedialog.asksaveasfilename(
            title="Export Configuration to Excel Workbook",
            defaultextension=".xlsx",
            filetypes=[("Excel Files (*.xlsx)", "*.xlsx"), ("All Files (*.*)", "*.*")],
            initialfile="fillable_pdf_configurator_backup.xlsx"
        )
        if not dest_file:
            return

        try:
            abs_dest = os.path.abspath(dest_file)
            self.save_all_sheets_to_excel(target_path=abs_dest)
            msg = f"SUCCESS! Configuration exported to Excel backup: '{abs_dest}'"
            self.set_inline_status(msg, is_success=True)
            self.log_info(msg)
        except PermissionError:
            self.set_inline_status(f"Permission Error: '{dest_file}' is currently open in Microsoft Excel. Please close Excel and try again.", is_error=True)
        except Exception as e:
            err_msg = f"Export Excel Error: {e}"
            self.set_inline_status(err_msg, is_error=True)
            self.log_error(f"{err_msg}\n{traceback.format_exc()}")

    def export_json_dialog(self):
        """Prompt user for destination file and export full configuration to JSON."""
        self.sync_ui_to_settings()
        dest_file = filedialog.asksaveasfilename(
            title="Export Configuration to JSON File",
            defaultextension=".json",
            filetypes=[("JSON Configuration (*.json)", "*.json"), ("All Files (*.*)", "*.*")],
            initialfile="fillable_pdf_configurator_backup.json"
        )
        if not dest_file:
            return

        try:
            abs_dest = os.path.abspath(dest_file)
            self.export_config_to_json(abs_dest)
            msg = f"SUCCESS! Configuration exported to JSON backup: '{abs_dest}'"
            self.set_inline_status(msg, is_success=True)
            self.log_info(msg)
        except Exception as e:
            err_msg = f"Export JSON Error: {e}"
            self.set_inline_status(err_msg, is_error=True)
            self.log_error(f"{err_msg}\n{traceback.format_exc()}")

    def download_ai_schema_dialog(self):
        """Prompt user to save/export the pure JSON schema specification template for supplying to outside AI models."""
        dest_file = filedialog.asksaveasfilename(
            title="Save Pure AI Form Configuration Schema Specification",
            defaultextension=".json",
            filetypes=[("JSON Schema Files (*.json)", "*.json"), ("All Files (*.*)", "*.*")],
            initialfile="ai_form_config_schema_template.json"
        )
        if not dest_file:
            return

        full_schema = {
            "$schema": "http://json-schema.org/draft-07/schema#",
            "title": "Interactive Fillable PDF Form Configuration Schema Specification",
            "description": "Supply this exact pure JSON schema contract to any AI model (ChatGPT, Claude, Gemini) to generate custom valid form configurations for the Interactive PDF Suite.",
            "type": "object",
            "required": ["form_settings", "section_rows", "field_schema_rows"],
            "properties": {
                "form_settings": {
                    "type": "object",
                    "description": "Global form basic details, theme colors, instructions, and summary section labels.",
                    "required": [
                        "Form_Title", "Form_Subtitle", "Instruction_Point_2",
                        "Footer_Text", "Default_Lock_Status", "Theme_Navy_Primary", "Theme_Soft_Blue",
                        "Theme_Soft_Yellow", "Summary_Section_Title", "Summary_Field_Label"
                    ],
                    "properties": {
                        "Form_Title": { "type": "string", "maxLength": 60, "description": "Form title displayed on PDF line 1" },
                        "Form_Subtitle": { "type": "string", "maxLength": 85, "description": "Form subtitle displayed on PDF line 2" },
                        "Instruction_Point_1": { "type": "string", "maxLength": 140, "description": "Optional in JSON. System-managed mandatory red save rule (auto-injected if omitted)" },
                        "Instruction_Point_2": { "type": "string", "maxLength": 110, "description": "Instruction point 2" },
                        "Instruction_Point_3": { "type": "string", "maxLength": 110, "description": "Instruction point 3 (Optional)" },
                        "Footer_Text": { "type": "string", "maxLength": 110, "description": "Page footer banner text" },
                        "Default_Lock_Status": { "type": "string", "enum": ["Locked", "Unlocked", "Not Required"] },
                        "Theme_Navy_Primary": { "type": "string", "pattern": "^#[0-9A-Fa-f]{6}$", "default": "#1F4E78" },
                        "Theme_Soft_Blue": { "type": "string", "pattern": "^#[0-9A-Fa-f]{6}$", "default": "#DDEBF7" },
                        "Theme_Soft_Yellow": { "type": "string", "pattern": "^#[0-9A-Fa-f]{6}$", "default": "#FFF2CC" },
                        "Summary_Section_Title": { "type": "string", "maxLength": 110, "description": "Summary section header title" },
                        "Summary_Button_Text": { "type": "string", "maxLength": 110, "default": ">>>> Click the Button To Generate Summary <<<<", "description": "Optional in JSON. System-managed button text (auto-injected if omitted)" },
                        "Summary_Field_Label": { "type": "string", "maxLength": 110, "description": "Summary output text label" }
                    }
                },
                "section_rows": {
                    "type": "array",
                    "description": "List of form layout section tuples: [Section_Number, Section_ID, Section_Heading_Title, Layout_Grid_Columns]",
                    "items": {
                        "type": "array",
                        "items": [
                            { "type": "integer", "description": "Section Number (1, 2, 3, 4...)" },
                            { "type": "string", "description": "Section ID (SEC_1, SEC_2, SEC_3...)" },
                            { "type": "string", "description": "Section Heading Title" },
                            { "type": "string", "enum": ["1", "2", "3", "4"], "description": "Layout Grid Columns (1, 2, 3, or 4)" }
                        ]
                    }
                },
                "field_schema_rows": {
                    "type": "array",
                    "description": "List of field definition tuples: [Field_ID, Section_ID, Field_Label, Field_DataType, Default_Value, Tooltip_Help_Text, Validation_Operator, Validation_Param_Value, Dropdown_Options]",
                    "items": {
                        "type": "array",
                        "items": [
                            { "type": "string", "description": "Field ID (Field_1_1, Field_1_2, Field_2_1...)" },
                            { "type": "string", "description": "Section ID matching a SEC_N defined in section_rows" },
                            { "type": "string", "description": "Field label heading displayed on PDF canvas" },
                            { "type": "string", "enum": ["short_text", "long_text", "whole_number", "decimal", "choice", "multi_choice", "radio_choice", "date"] },
                            { "type": "string", "description": "Default pre-populated value" },
                            { "type": "string", "description": "Tooltip help text shown on hover" },
                            { "type": "string", "enum": ["none", "min_length", "max_length", "=", "<", ">", "<=", ">=", "in_range", "out_range", "date_ddmmyyyy"] },
                            { "type": "string", "description": "Validation parameter value (Required for min_length, max_length, in_range, comparison operators)" },
                            { "type": "string", "description": "Comma-separated list of dropdown choices (Only required when Field_DataType is 'choice')" }
                        ]
                    }
                }
            }
        }

        try:
            abs_dest = os.path.abspath(dest_file)
            with open(abs_dest, "w", encoding="utf-8") as f_out:
                json.dump(full_schema, f_out, indent=4)

            msg = f"SUCCESS! Pure AI Schema Specification saved to: '{abs_dest}'"
            self.set_inline_status(msg, is_success=True)
            self.log_info(msg)
            try:
                os.startfile(abs_dest)
            except Exception:
                pass
        except Exception as e:
            err_msg = f"Download AI Schema Error: {e}"
            self.set_inline_status(err_msg, is_error=True)
            self.log_error(f"{err_msg}\n{traceback.format_exc()}")

    def export_config_to_json(self, json_path: str):
        """Exports the full 3-sheet configuration state to a clean JSON file."""
        data = {
            "form_settings": self.form_settings,
            "section_rows": self.section_rows,
            "field_schema_rows": self.field_schema_rows
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)

    def import_config_dialog(self):
        """Prompt user to select a JSON or Excel configurator file and import its schema into the GUI."""
        src_file = filedialog.askopenfilename(
            title="Import Configuration (JSON or Excel)",
            filetypes=[("Supported Config Files", "*.json;*.xlsx"), ("JSON Configuration", "*.json"), ("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if not src_file:
            return

        abs_src = os.path.abspath(src_file)
        try:
            if abs_src.lower().endswith(".json"):
                self.load_state_from_json_file(abs_src)
                msg = f"SUCCESS! Imported configuration from JSON: '{abs_src}'"
            else:
                self.load_state_from_excel_file(abs_src)
                msg = f"SUCCESS! Imported configuration from Excel: '{abs_src}'"
            self.set_inline_status(msg, is_success=True)
            self.log_info(msg)
        except Exception as e:
            err_msg = f"Import Error: Could not read configuration from '{abs_src}': {e}"
            self.set_inline_status(err_msg, is_error=True)
            self.log_error(f"{err_msg}\n{traceback.format_exc()}")

    def load_state_from_json_file(self, json_path: str):
        """Loads configuration state from a JSON file and refreshes all UI tabs."""
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if "form_settings" in data and isinstance(data["form_settings"], dict):
            self.form_settings.update(data["form_settings"])
            if hasattr(self, 'ent_title'):
                self.ent_title.delete(0, tk.END); self.ent_title.insert(0, self.form_settings.get("Form_Title", ""))
                self.ent_subtitle.delete(0, tk.END); self.ent_subtitle.insert(0, self.form_settings.get("Form_Subtitle", ""))
                self.ent_inst1.configure(state="normal")
                self.ent_inst1.delete(0, tk.END); self.ent_inst1.insert(0, "!! Save As - this PDF file before mailing. Cross-check whether all filled form data is properly saved before sending. !!")
                self.ent_inst1.configure(state="disabled")
                self.ent_inst2.configure(state="normal")
                self.ent_inst2.delete(0, tk.END); self.ent_inst2.insert(0, "!! Before saving the file Click 'Generate Summary' Button to update it..!!   !! Use Adobe Acrobat Reader software for filling this form. !!")
                self.ent_inst2.configure(state="disabled")
                self.ent_inst3.delete(0, tk.END); self.ent_inst3.insert(0, self.form_settings.get("Instruction_Point_3", ""))
                if hasattr(self, 'ent_inst4'):
                    self.ent_inst4.delete(0, tk.END); self.ent_inst4.insert(0, self.form_settings.get("Instruction_Point_4", ""))
                self.ent_footer.delete(0, tk.END); self.ent_footer.insert(0, self.form_settings.get("Footer_Text", ""))
                self.cmb_lock.set(self.form_settings.get("Default_Lock_Status", "Locked"))
            if hasattr(self, 'ent_sum_sec_title'):
                self.ent_sum_sec_title.configure(state="normal")
                self.ent_sum_sec_title.delete(0, tk.END); self.ent_sum_sec_title.insert(0, "GENERATE FORM SUMMARY & COPY TEXT")
                self.ent_sum_sec_title.configure(state="disabled")
                self.ent_sum_btn_txt.configure(state="normal")
                self.ent_sum_btn_txt.delete(0, tk.END); self.ent_sum_btn_txt.insert(0, ">>>> Click the Button - To Generate Form Summary <<<<")
                self.ent_sum_btn_txt.configure(state="disabled")
                self.ent_sum_fld_lbl.delete(0, tk.END); self.ent_sum_fld_lbl.insert(0, self.form_settings.get("Summary_Field_Label", ""))

        if "section_rows" in data and isinstance(data["section_rows"], list):
            secs = []
            for row in data["section_rows"]:
                if row and len(row) >= 4:
                    s_id = str(row[1] or '').strip()
                    s_title = str(row[2] or '').strip()
                    if s_id.upper() == "SEC_5" or "GENERATE SHORT DPR" in s_title.upper() or "SHORT DPR SUMMARY" in s_title.upper():
                        continue
                    secs.append([row[0], s_id, s_title, str(row[3] or '3').strip()])
            if secs:
                self.section_rows = secs
                self.refresh_sections_treeview()

        if "field_schema_rows" in data and isinstance(data["field_schema_rows"], list):
            flds = []
            for row in data["field_schema_rows"]:
                if row and len(row) >= 8:
                    f_id = str(row[0]).strip()
                    sec_id = str(row[1] or '').strip()
                    f_lbl = str(row[2] or '').strip()
                    if f_id.upper() == "FIELD_5_1" or sec_id.upper() == "SEC_5" or "GENERATED SHORT DPR" in f_lbl.upper():
                        continue
                    def_val = str(row[4] or '').strip()
                    opts_val = str(row[8] or '').strip() if len(row) > 8 else ''


                    if f_id.upper() == "FIELD_4_1" and "ENTER PROGRESS SUMMARY" in f_lbl.upper():
                        f_lbl = "OPERATIONAL PROGRESS NARRATIVE"

                    flds.append([
                        f_id, sec_id, f_lbl,
                        str(row[3] or 'short_text').strip().lower(), def_val,
                        str(row[5] or '').strip(), str(row[6] or 'none').strip().lower(),
                        str(row[7] or '').strip(), opts_val
                    ])
            if flds:
                self.field_schema_rows = flds
                self.refresh_fields_treeview()

        self.save_all_sheets_to_excel()

    def save_all_sheets_to_excel(self, target_path=None):
        """Synchronizes all 3 sheets from Tkinter UI state to specified Excel workbook."""
        save_path = target_path or self.excel_path
        create_3sheet_pdf_configurator_excel(save_path)
        wb = openpyxl.load_workbook(save_path)

        font_bold_navy = Font(name="Calibri", size=10, bold=True, color="1F4E78")
        font_regular = Font(name="Calibri", size=10, color="000000")
        fill_editable = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        border_thin = Border(
            left=Side(style='thin', color='B0C4DE'), right=Side(style='thin', color='B0C4DE'),
            top=Side(style='thin', color='B0C4DE'), bottom=Side(style='thin', color='B0C4DE')
        )
        prot_unlocked = Protection(locked=False)
        prot_locked = Protection(locked=True)

        # Sheet 1: Form_Basic_Details
        ws1 = wb["Form_Basic_Details"]
        for row in ws1.iter_rows(min_row=2):
            if row and row[0].value:
                key = str(row[0].value).strip()
                if key in self.form_settings:
                    row[1].value = self.form_settings[key]

        # Sheet 2: Section_Details
        ws2 = wb["Section_Details"]
        for row_idx in range(ws2.max_row, 1, -1):
            ws2.delete_rows(row_idx)

        for idx, sec in enumerate(self.section_rows, start=2):
            ws2.append(sec)
            for c_idx in range(1, 5):
                cell = ws2.cell(row=idx, column=c_idx)
                cell.border = border_thin
                if c_idx in (1, 2):
                    cell.font = font_bold_navy
                    cell.protection = prot_locked
                    if idx % 2 == 1: cell.fill = fill_zebra
                else:
                    cell.font = font_bold_navy if c_idx == 3 else font_regular
                    cell.fill = fill_editable
                    cell.protection = prot_unlocked

        # Sheet 3: Field_Configuration
        ws3 = wb["Field_Configuration"]
        for row_idx in range(ws3.max_row, 1, -1):
            ws3.delete_rows(row_idx)

        for idx, fld in enumerate(self.field_schema_rows, start=2):
            ws3.append(fld)
            f_dtype = str(fld[3] or '').strip().lower()
            val_op = str(fld[6] or 'none').strip().lower()

            for c_idx in range(1, 10):
                cell = ws3.cell(row=idx, column=c_idx)
                cell.border = border_thin
                if c_idx in (1, 2):
                    cell.font = font_bold_navy
                    cell.protection = prot_locked
                    if idx % 2 == 1: cell.fill = fill_zebra
                elif c_idx == 8:
                    requires_param = val_op not in ('none', 'date_ddmmyyyy', '')
                    if requires_param:
                        cell.font = font_regular
                        cell.fill = fill_editable
                        cell.protection = prot_unlocked
                    else:
                        cell.font = font_regular
                        cell.fill = fill_zebra
                        cell.protection = prot_locked
                elif c_idx == 9:
                    if f_dtype == 'choice':
                        cell.font = font_regular
                        cell.fill = fill_editable
                        cell.protection = prot_unlocked
                    else:
                        cell.font = font_regular
                        cell.fill = fill_zebra
                        cell.protection = prot_locked
                else:
                    cell.font = font_regular
                    cell.fill = fill_editable
                    cell.protection = prot_unlocked

        wb.save(save_path)

    def load_initial_state_from_excel(self):
        """Loads configuration from master Excel file if it exists at startup."""
        if os.path.exists(self.excel_path):
            try:
                self.load_state_from_excel_file(self.excel_path)
            except Exception as e:
                self.log_error(f"Error loading initial state from Excel: {e}")

    def load_state_from_excel_file(self, target_path: str):
        """Loads configuration from specified 3-sheet Excel file into the GUI state."""
        if not os.path.exists(target_path):
            raise FileNotFoundError(f"Specified Excel file does not exist: '{target_path}'")

        wb = openpyxl.load_workbook(target_path, data_only=True)
        
        # Sheet 1: Form_Basic_Details
        if "Form_Basic_Details" in wb.sheetnames:
            ws1 = wb["Form_Basic_Details"]
            for row in ws1.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    k = str(row[0]).strip()
                    v = str(row[1] or '').strip()
                    if k in self.form_settings:
                        self.form_settings[k] = v

            # Refresh Tab 1 Entry Widgets
            if hasattr(self, 'ent_title'):
                self.ent_title.delete(0, tk.END); self.ent_title.insert(0, self.form_settings.get("Form_Title", ""))
                self.ent_subtitle.delete(0, tk.END); self.ent_subtitle.insert(0, self.form_settings.get("Form_Subtitle", ""))
                self.ent_inst1.configure(state="normal")
                self.ent_inst1.delete(0, tk.END); self.ent_inst1.insert(0, "!! Save As - this PDF file before mailing. Cross-check whether all filled form data is properly saved before sending. !!")
                self.ent_inst1.configure(state="disabled")
                self.ent_inst2.configure(state="normal")
                self.ent_inst2.delete(0, tk.END); self.ent_inst2.insert(0, "!! Before saving the file Click 'Generate Summary' Button to update it..!!   !! Use Adobe Acrobat Reader software for filling this form. !!")
                self.ent_inst2.configure(state="disabled")
                self.ent_inst3.delete(0, tk.END); self.ent_inst3.insert(0, self.form_settings.get("Instruction_Point_3", ""))
                if hasattr(self, 'ent_inst4'):
                    self.ent_inst4.delete(0, tk.END); self.ent_inst4.insert(0, self.form_settings.get("Instruction_Point_4", ""))
                self.ent_footer.delete(0, tk.END); self.ent_footer.insert(0, self.form_settings.get("Footer_Text", ""))
                self.cmb_lock.set(self.form_settings.get("Default_Lock_Status", "Locked"))
            if hasattr(self, 'ent_sum_sec_title'):
                self.ent_sum_sec_title.configure(state="normal")
                self.ent_sum_sec_title.delete(0, tk.END); self.ent_sum_sec_title.insert(0, "GENERATE FORM SUMMARY & COPY TEXT")
                self.ent_sum_sec_title.configure(state="disabled")
                self.ent_sum_btn_txt.configure(state="normal")
                self.ent_sum_btn_txt.delete(0, tk.END); self.ent_sum_btn_txt.insert(0, ">>>> Click the Button - To Generate Form Summary <<<<")
                self.ent_sum_btn_txt.configure(state="disabled")
                self.ent_sum_fld_lbl.delete(0, tk.END); self.ent_sum_fld_lbl.insert(0, self.form_settings.get("Summary_Field_Label", ""))

        # Sheet 2: Section_Details
        ws2 = wb["Section_Details"] if "Section_Details" in wb.sheetnames else None
        if ws2:
            secs = []
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    s_id = str(row[1] or '').strip()
                    s_title = str(row[2] or '').strip()
                    # Filter out legacy hardcoded summary section rows
                    if s_id.upper() == "SEC_5" or "GENERATE SHORT DPR" in s_title.upper() or "SHORT DPR SUMMARY" in s_title.upper():
                        continue
                    secs.append([row[0], s_id, s_title, str(row[3] or '3').strip()])
            if secs:
                self.section_rows = secs
                self.refresh_sections_treeview()

        # Sheet 3: Field_Configuration
        ws3 = wb["Field_Configuration"] if "Field_Configuration" in wb.sheetnames else None
        if ws3:
            flds = []
            for row in ws3.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    f_id = str(row[0]).strip()
                    sec_id = str(row[1] or '').strip()
                    f_lbl = str(row[2] or '').strip()
                    # Filter out legacy hardcoded summary field rows
                    if f_id.upper() == "FIELD_5_1" or sec_id.upper() == "SEC_5" or "GENERATED SHORT DPR" in f_lbl.upper():
                        continue
                    
                    def_val = str(row[4] or '').strip()
                    opts_val = str(row[8] or '').strip() if len(row) > 8 else ''



                    if f_id.upper() == "FIELD_4_1" and "ENTER PROGRESS SUMMARY" in f_lbl.upper():
                        f_lbl = "OPERATIONAL PROGRESS NARRATIVE"

                    flds.append([
                        f_id, sec_id, f_lbl,
                        str(row[3] or 'short_text').strip().lower(), def_val,
                        str(row[5] or '').strip(), str(row[6] or 'none').strip().lower(),
                        str(row[7] or '').strip(), opts_val
                    ])
            if flds:
                self.field_schema_rows = flds
                self.refresh_fields_treeview()

    def browse_pdf_folder(self):
        folder = filedialog.askdirectory(initialdir=self.output_pdf_folder)
        if folder:
            self.ent_pdf_dir.delete(0, tk.END)
            self.ent_pdf_dir.insert(0, os.path.abspath(folder))

    def run_data_consolidation(self):
        target_dir = self.ent_pdf_dir.get().strip()
        os.makedirs(target_dir, exist_ok=True)  # Auto-create directory if it does not exist

        self.set_inline_status("Running batch PDF AcroForm data extraction...")
        try:
            out_excel, pdf_count = consolidate_pdf_forms_to_excel(target_dir)
            abs_excel = os.path.abspath(out_excel)

            if pdf_count == 0:
                self.set_inline_status(f"No PDF files were found in: '{target_dir}'", is_error=True)
                self.lbl_consol_result.configure(text=f"⚠️ No PDF files found in '{target_dir}'.", fg="#C00000")
                return

            msg = f"SUCCESS! Consolidated data from {pdf_count} PDF forms into Excel: '{abs_excel}'"
            self.set_inline_status(msg, is_success=True)
            self.log_info(msg)
            self.lbl_consol_result.configure(text=f"✔ Consolidated {pdf_count} PDFs into Excel report: {abs_excel}", fg="#006100")

            try:
                os.startfile(abs_excel)
            except Exception:
                pass

        except PermissionError:
            self.set_inline_status("File Permission Error: Consolidated Excel report is currently open in Microsoft Excel! Please close Excel and try again.", is_error=True)
        except Exception as e:
            err_msg = f"Consolidation Error: {e}"
            self.set_inline_status(err_msg, is_error=True)
            self.log_error(f"{err_msg}\n{traceback.format_exc()}")

    def open_log_file(self):
        if os.path.exists(LOG_FILE_PATH):
            os.startfile(LOG_FILE_PATH)


def main():
    app = DigitalDPRSuiteGUI()
    app.mainloop()


if __name__ == "__main__":
    main()
