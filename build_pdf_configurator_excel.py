"""
Master 3-Sheet Excel Form Configurator Builder Module
------------------------------------------------------
Creates the master Fillable PDF Form Configurator Excel workbook:
'Interactive_PDF_Code/fillable_pdf_configurator.xlsx'

Directory Architecture:
- Located in 'Interactive_PDF_Code/'.
- Outputs master Excel file: 'Interactive_PDF_Code/fillable_pdf_configurator.xlsx'.
- Generates clean 3-sheet structure (Form_Basic_Details, Section_Details, Field_Configuration).
- Contextual Cell Protection, Cascading Data Validation (=INDIRECT(D2)), Range Boundary Validation, and Clean 9-Column Schema.
- Robust File Lock Handling (PermissionError) and Schema Validation checks.

Follows project coding standards (Python 3.10+ types, openpyxl, logging).
"""

import os
import logging
import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

logger = logging.getLogger(__name__)


def validate_excel_config(excel_path: str = "Interactive_PDF_Code/fillable_pdf_configurator.xlsx") -> tuple[bool, str]:
    """
    Validates user entries inside the 3-sheet Excel Configurator workbook.

    Returns:
        tuple[bool, str]: (is_valid, error_description)
    """
    abs_path = os.path.abspath(excel_path)
    if not os.path.exists(abs_path):
        return False, f"Master Excel Configurator file not found at: '{abs_path}'"

    try:
        wb = openpyxl.load_workbook(abs_path, data_only=True)
    except PermissionError:
        return False, f"File Permission Error: '{abs_path}' is currently open in Microsoft Excel! Please close Microsoft Excel and try again."
    except Exception as e:
        return False, f"Could not read Excel file: {e}"

    # Check Sheet 1: Form_Basic_Details
    cfg_sheet_name = "Form_Basic_Details" if "Form_Basic_Details" in wb.sheetnames else ("Form_Config" if "Form_Config" in wb.sheetnames else None)
    if not cfg_sheet_name:
        return False, "Missing mandatory Sheet 1 'Form_Basic_Details' in Excel workbook."

    ws_cfg = wb[cfg_sheet_name]
    cfg_dict = {}
    for row in ws_cfg.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            cfg_dict[str(row[0]).strip()] = str(row[1] or '').strip()

    title = cfg_dict.get('Form_Title', '')
    if not title:
        return False, "Validation Error in Sheet 1: 'Form_Title' cannot be empty!"
    if len(title) > 60:
        return False, f"Validation Error in Sheet 1: 'Form_Title' length ({len(title)} chars) exceeds max limit of 60 characters."

    subtitle = cfg_dict.get('Form_Subtitle', '')
    if len(subtitle) > 85:
        return False, f"Validation Error in Sheet 1: 'Form_Subtitle' length ({len(subtitle)} chars) exceeds max limit of 85 characters."

    # Check Theme Color Codes format
    for color_key in ['Theme_Navy_Primary', 'Theme_Soft_Blue', 'Theme_Soft_Yellow']:
        c_code = cfg_dict.get(color_key, '')
        if c_code and (len(c_code) != 7 or not c_code.startswith('#')):
            return False, f"Validation Error in Sheet 1: '{color_key}' value '{c_code}' is not a valid 7-character HTML Hex color code (e.g. #1F4E78)."

    # Check Sheet 2: Section_Details
    if "Section_Details" not in wb.sheetnames:
        return False, "Missing mandatory Sheet 2 'Section_Details' in Excel workbook."

    # Check Sheet 3: Field_Configuration
    sch_sheet_name = "Field_Configuration" if "Field_Configuration" in wb.sheetnames else ("Field_Schema" if "Field_Schema" in wb.sheetnames else None)
    if not sch_sheet_name:
        return False, "Missing mandatory Sheet 3 'Field_Configuration' in Excel workbook."

    ws_sch = wb[sch_sheet_name]
    field_count = 0
    for row in ws_sch.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            field_count += 1
            f_dtype = str(row[3] or '').strip().lower()
            val_op = str(row[6] or 'none').strip().lower()
            val_param = str(row[7] or '').strip()
            if f_dtype and f_dtype not in ('short_text', 'long_text', 'whole_number', 'decimal', 'choice', 'multi_choice', 'radio_choice', 'date'):
                return False, f"Validation Error in Row {field_count + 1} (Field_ID '{row[0]}'): Invalid Field_DataType '{f_dtype}'. Must be short_text, long_text, whole_number, decimal, choice, multi_choice, radio_choice, or date."

            # Check range boundary contract
            if val_op in ('in_range', 'out_range'):
                parts = [p.strip() for p in val_param.split(',') if p.strip()]
                if len(parts) != 2:
                    return False, f"Validation Error in Row {field_count + 1} (Field_ID '{row[0]}'): Operator '{val_op}' requires 2 numbers separated by a comma (e.g. '18, 60'). Got: '{val_param}'"
                try:
                    p1, p2 = float(parts[0]), float(parts[1])
                    if p1 >= p2:
                        return False, f"Validation Error in Row {field_count + 1} (Field_ID '{row[0]}'): Lower bound ({p1}) must be strictly smaller than upper bound ({p2})."
                except ValueError:
                    return False, f"Validation Error in Row {field_count + 1} (Field_ID '{row[0]}'): Range bounds must be valid numbers (e.g. '18, 60'). Got: '{val_param}'"

    if field_count == 0:
        return False, "Validation Error in Sheet 3: No field configurations found in 'Field_Configuration'."

    return True, "Excel Configuration Validated Successfully!"


def create_3sheet_pdf_configurator_excel(output_excel_path: str = "Interactive_PDF_Code/fillable_pdf_configurator.xlsx") -> str:
    """
    Creates the master 3-sheet Excel Configurator workbook in 'Interactive_PDF_Code/fillable_pdf_configurator.xlsx'.
    Handles PermissionError gracefully if the file is currently open in Microsoft Excel.
    """
    abs_path = os.path.abspath(output_excel_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)

    wb = openpyxl.Workbook()

    # Typography & Styling Tokens
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_bold_navy = Font(name="Calibri", size=10, bold=True, color="1F4E78")
    font_regular = Font(name="Calibri", size=10, color="000000")
    
    # Soft Blue Fill (#DDEBF7) for User Editable Input Cells
    fill_editable = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border_thin = Border(
        left=Side(style='thin', color='B0C4DE'),
        right=Side(style='thin', color='B0C4DE'),
        top=Side(style='thin', color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE')
    )

    prot_unlocked = Protection(locked=False)
    prot_locked = Protection(locked=True)

    # =========================================================================
    # SHEET 1: Form_Basic_Details (Clean 3-Column Compiler Structure)
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Form_Basic_Details"
    ws1.views.sheetView[0].showGridLines = True

    ws1.append(["Setting_Key (Fixed Key)", "Setting_Value (Editable)", "Help_Prompt / Description"])

    config_rows = [
        ["Form_Title", "Employee_Information_Form", "Header Banner Line 1 Title (Max 60 characters for 14pt Bold)"],
        ["Form_Subtitle", "OFFICIAL EMPLOYEE DETAILS & RECORD FORM", "Header Banner Line 2 Subtitle (Max 85 characters for 10pt Bold)"],
        ["Instruction_Point_1", "!! Save As - this PDF file before mailing. Cross-check whether all filled form data is properly saved before sending. !!", "Mandatory System Instruction Point 1 [Locked System Rule]"],
        ["Instruction_Point_2", "!! Before saving the file Click 'Generate Summary' Button to update it..!!   !! Use Adobe Acrobat Reader software for filling this form. !!", "Mandatory System Instruction Point 2 [Locked System Rule]"],
        ["Instruction_Point_3", "3. Verify all employee personal details & department info before generating summary.", "Instruction Point 3 (User Configurable - Max 110 characters)"],
        ["Instruction_Point_4", "", "Instruction Point 4 (Optional User Configurable - Max 110 characters)"],
        ["Footer_Text", "Official Employee Information & Department Record Form", "Page Footer Banner Text (Max 110 characters for 8pt Oblique)"],
        ["Default_Lock_Status", "Unlocked", "Initial Lock Status: Select 'Locked', 'Unlocked', or 'Not Required' (removes lock button)"],
        ["Theme_Navy_Primary", "#1F4E78", "Primary Header Banner & Border Color Code (e.g. #1F4E78)"],
        ["Theme_Soft_Blue", "#DDEBF7", "Instruction Box Container Fill Color Code (e.g. #DDEBF7)"],
        ["Theme_Soft_Yellow", "#FFF2CC", "Summary Box Container Fill Color Code (e.g. #FFF2CC)"],
        ["Summary_Section_Title", "GENERATE FORM SUMMARY & COPY TEXT", "Summary Section Title [Locked System Text]"],
        ["Summary_Button_Text", ">>>> Click the Button - To Generate Form Summary <<<<", "Mandatory System Action Button Text [Locked System Text]"],
        ["Summary_Field_Label", "GENERATED FORM SUMMARY TEXT", "Summary Output Text Field Label (Displayed above Output Box)"]
    ]

    for r in config_rows:
        ws1.append(r)

    # Format Header Row for Sheet 1
    for col_idx in range(1, 4):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.protection = prot_locked

    font_bold_red = Font(name="Calibri", size=10, bold=True, color="C00000")
    fill_locked_gray = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")

    for row_idx in range(2, len(config_rows) + 2):
        cell_key = ws1.cell(row=row_idx, column=1)
        cell_val = ws1.cell(row=row_idx, column=2)
        cell_desc = ws1.cell(row=row_idx, column=3)

        # Setting_Key -> Locked (Fixed Options)
        cell_key.font = font_bold_navy
        cell_key.border = border_thin
        cell_key.protection = prot_locked

        key_name = str(cell_key.value or "").strip()

        if key_name in ("Instruction_Point_1", "Instruction_Point_2", "Summary_Section_Title", "Summary_Button_Text"):
            # SYSTEM-LOCKED SETTING VALUES -> Locked, Non-editable, Grayed Out
            cell_val.font = font_bold_red if key_name.startswith("Instruction") else font_bold_navy
            cell_val.border = border_thin
            cell_val.fill = fill_locked_gray
            cell_val.protection = prot_locked
        else:
            # User Editable Setting Values -> Soft Blue Fill, Unlocked
            cell_val.font = font_regular
            cell_val.border = border_thin
            cell_val.fill = fill_editable
            cell_val.protection = prot_unlocked

        # Help_Prompt / Description -> Locked
        cell_desc.font = font_regular
        cell_desc.border = border_thin
        cell_desc.protection = prot_locked

    # Turn ON Sheet Protection for Sheet 1
    ws1.protection.sheet = True

    # --- SHEET 1 EXCEL DATA VALIDATIONS (EXPLICIT SHOW ERROR MESSAGE = TRUE) ---
    
    # 1. Form_Title (Max 60 Chars for 14pt Bold Font) -> Cell B2
    dv_title = DataValidation(
        type="textLength", operator="lessThanOrEqual", formula1="60", allow_blank=False,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Character Limit Exceeded!",
        error="Max characters allowed for Form_Title is 60 characters. Please shorten your text to fit the 14pt Bold header banner.",
        promptTitle="Character Limit", prompt="Max 60 characters allowed for Form_Title."
    )
    ws1.add_data_validation(dv_title)
    dv_title.add("B2")

    # 2. Form_Subtitle (Max 85 Chars for 10pt Bold Font) -> Cell B3
    dv_subtitle = DataValidation(
        type="textLength", operator="lessThanOrEqual", formula1="85", allow_blank=False,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Character Limit Exceeded!",
        error="Max characters allowed for Form_Subtitle is 85 characters. Please shorten your text to fit the 10pt Bold subtitle.",
        promptTitle="Character Limit", prompt="Max 85 characters allowed for Form_Subtitle."
    )
    ws1.add_data_validation(dv_subtitle)
    dv_subtitle.add("B3")

    # 3. Instruction_Point_3 (Max 110 Chars) -> Cell B6
    dv_inst3 = DataValidation(
        type="textLength", operator="lessThanOrEqual", formula1="110", allow_blank=True,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Character Limit Exceeded!",
        error="Max characters allowed for Instruction_Point_3 is 110 characters.",
        promptTitle="Character Limit", prompt="Max 110 characters allowed for Instruction_Point_3."
    )
    ws1.add_data_validation(dv_inst3)
    dv_inst3.add("B6")

    # 4. Instruction_Point_4 (Max 110 Chars) -> Cell B7
    dv_inst4 = DataValidation(
        type="textLength", operator="lessThanOrEqual", formula1="110", allow_blank=True,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Character Limit Exceeded!",
        error="Max characters allowed for Instruction_Point_4 is 110 characters.",
        promptTitle="Character Limit", prompt="Max 110 characters allowed for Instruction_Point_4."
    )
    ws1.add_data_validation(dv_inst4)
    dv_inst4.add("B7")

    # 5. Footer_Text (Max 110 Chars for 8pt Oblique Font) -> Cell B8
    dv_footer = DataValidation(
        type="textLength", operator="lessThanOrEqual", formula1="110", allow_blank=True,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Character Limit Exceeded!",
        error="Max characters allowed for Footer_Text is 110 characters.",
        promptTitle="Character Limit", prompt="Max 110 characters allowed for Footer_Text."
    )
    ws1.add_data_validation(dv_footer)
    dv_footer.add("B8")

    # 7. Default_Lock_Status In-Cell Dropdown List -> Cell B9
    dv_lock = DataValidation(
        type="list", formula1='"Locked, Unlocked, Not Required"', allow_blank=False,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Invalid Lock Status!",
        error="Invalid Lock Status! Please select 'Locked', 'Unlocked', or 'Not Required' from the dropdown list.",
        promptTitle="Initial Lock Status", prompt="Select 'Locked', 'Unlocked', or 'Not Required'."
    )
    ws1.add_data_validation(dv_lock)
    dv_lock.add("B9")

    # 8. HTML Color Code Hex Validation -> Cells B10:B12
    dv_color = DataValidation(
        type="custom",
        formula1='AND(LEN(B10)=7, LEFT(B10,1)="#")',
        allow_blank=False,
        showErrorMessage=True,
        showInputMessage=True,
        errorTitle="Invalid HTML Color Code Format!",
        error="Non-standard color code entered! Please enter a valid 7-character HTML Hex Color Code starting with # (e.g. #1F4E78, #DDEBF7, #FFF2CC).",
        promptTitle="HTML Hex Color Code",
        prompt="Enter 7-character HTML Hex Color Code starting with # (e.g. #1F4E78)."
    )
    ws1.add_data_validation(dv_color)
    dv_color.add("B10:B12")

    # 9. Summary Settings Character Length Validation -> Cells B13:B15 (Max 110 Chars)
    dv_sum = DataValidation(
        type="textLength", operator="lessThanOrEqual", formula1="110", allow_blank=False,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Character Limit Exceeded!",
        error="Max characters allowed for Summary Settings is 110 characters.",
        promptTitle="Character Limit", prompt="Max 110 characters allowed."
    )
    ws1.add_data_validation(dv_sum)
    dv_sum.add("B13:B15")

    # Turn ON Sheet Protection for Sheet 1
    ws1.protection.sheet = True

    # Auto-fit columns Sheet 1
    for col in ws1.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # =========================================================================
    # SHEET 2: Section_Details (Clean 4-Column Domain-Agnostic Structure)
    # =========================================================================
    ws2 = wb.create_sheet(title="Section_Details")
    ws2.views.sheetView[0].showGridLines = True

    ws2.append(["Section_Number", "Section_ID", "Section_Heading_Title", "Layout_Grid_Columns (Select 1, 2, 3, 4)"])

    section_rows = [
        [1, "SEC_1", "1. EMPLOYEE PERSONAL DETAILS", "3"],
        [2, "SEC_2", "2. DEPARTMENT & POSITION DETAILS", "3"],
        [3, "SEC_3", "3. RESIDENTIAL ADDRESS & REMARKS", "1"]
    ]

    for r in section_rows:
        ws2.append(r)

    # Format Header Row Sheet 2
    for col_idx in range(1, 5):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.protection = prot_locked

    for row_idx in range(2, len(section_rows) + 2):
        for col_idx in range(1, 5):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            
            # Structural Columns (Section_Number, Section_ID) -> Locked
            if col_idx in (1, 2):
                cell.font = font_bold_navy
                cell.protection = prot_locked
                if row_idx % 2 == 1:
                    cell.fill = fill_zebra
            else:
                # User Editable Columns (Section_Heading_Title, Layout_Grid_Columns) -> Soft Blue Fill & Unlocked
                cell.font = font_bold_navy if col_idx == 3 else font_regular
                cell.fill = fill_editable
                cell.protection = prot_unlocked
                if col_idx == 4:
                    cell.alignment = align_center

    # EXCEL IN-CELL DATA VALIDATION DROPDOWN FOR SECTION GRID COLUMNS (1, 2, 3, 4)
    dv_sec_col = DataValidation(
        type="list", formula1='"1, 2, 3, 4"', allow_blank=False,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Invalid Grid Column Count!",
        error="Invalid Column Count! Please select 1, 2, 3, or 4 from the dropdown list.",
        promptTitle="Grid Columns", prompt="Select 1, 2, 3, or 4 grid columns for this section layout."
    )
    ws2.add_data_validation(dv_sec_col)
    dv_sec_col.add(f"D2:D{len(section_rows) + 1}")

    # Turn ON Sheet Protection for Sheet 2
    ws2.protection.sheet = True

    # Auto-fit columns Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # =========================================================================
    # LOOKUP SHEET: _Validation_Lists (FOR EXCEL DEPENDENT CASCADING DROPDOWNS)
    # =========================================================================
    ws_val_lists = wb.create_sheet(title="_Validation_Lists")
    ws_val_lists.views.sheetView[0].showGridLines = True

    # Column mapping for Field_DataTypes:
    # A: short_text | B: long_text | C: whole_number | D: decimal | E: choice | F: date | G: multi_choice | H: radio_choice
    ws_val_lists.append(["short_text", "long_text", "whole_number", "decimal", "choice", "date", "multi_choice", "radio_choice"])

    val_matrix = [
        ["none", "none", "none", "none", "none", "none", "none", "none"],
        ["min_length", "min_length", "=", "=", "", "date_ddmmyyyy", "", ""],
        ["max_length", "max_length", "<", "<", "", "=", "", ""],
        ["", "", ">", ">", "", "<", "", ""],
        ["", "", "<=", "<=", "", ">", "", ""],
        ["", "", ">=", ">=", "", "<=", "", ""],
        ["", "", "in_range", "in_range", "", "in_range", "", ""],
        ["", "", "out_range", "out_range", "", "out_range", "", ""]
    ]

    for row in val_matrix:
        ws_val_lists.append(row)

    # Define openpyxl Named Ranges for Excel =INDIRECT(D2) Cascading Validation
    wb.defined_names.add(DefinedName("short_text", attr_text="'_Validation_Lists'!$A$2:$A$4"))
    wb.defined_names.add(DefinedName("long_text", attr_text="'_Validation_Lists'!$B$2:$B$4"))
    wb.defined_names.add(DefinedName("whole_number", attr_text="'_Validation_Lists'!$C$2:$C$9"))
    wb.defined_names.add(DefinedName("decimal", attr_text="'_Validation_Lists'!$D$2:$D$9"))
    wb.defined_names.add(DefinedName("choice", attr_text="'_Validation_Lists'!$E$2:$E$2"))
    wb.defined_names.add(DefinedName("date", attr_text="'_Validation_Lists'!$F$2:$F$9"))
    wb.defined_names.add(DefinedName("multi_choice", attr_text="'_Validation_Lists'!$G$2:$G$2"))
    wb.defined_names.add(DefinedName("radio_choice", attr_text="'_Validation_Lists'!$H$2:$H$2"))

    ws_val_lists.protection.sheet = True
    ws_val_lists.sheet_state = 'hidden'

    # =========================================================================
    # SHEET 3: Field_Configuration (STREAMLINED 9-COLUMN SCHEMA)
    # =========================================================================
    ws3 = wb.create_sheet(title="Field_Configuration")
    ws3.views.sheetView[0].showGridLines = True

    schema_headers = [
        "Field_ID (Locked)", "Section_ID (Locked)", "Field_Label", "Field_DataType (Dropdown)",
        "Default_Value", "Tooltip_Help_Text", "Validation_Operator (Conditional Dropdown)", "Validation_Param_Value (Conditional Input)",
        "Dropdown_Options (Conditional: Comma-Separated for choice datatype only)"
    ]
    ws3.append(schema_headers)

    # STREAMLINED SCHEMA ROWS WITH MIN_LENGTH & MAX_LENGTH VALIDATIONS
    schema_rows = [
        ["Field_1_1", "SEC_1", "FULL NAME", "short_text", "", "Enter employee full legal name", "max_length", "100", ""],
        ["Field_1_2", "SEC_1", "EMPLOYEE ID / BADGE NO", "short_text", "", "Enter official Employee ID or Badge number", "max_length", "50", ""],
        ["Field_1_3", "SEC_1", "DATE OF BIRTH (DD-MM-YYYY)", "date", "", "Enter date of birth in DD-MM-YYYY format", "date_ddmmyyyy", "", ""],
        ["Field_1_4", "SEC_1", "CONTACT PHONE NUMBER", "whole_number", "", "Enter 10-digit mobile contact number", ">=", "0", ""],
        ["Field_1_5", "SEC_1", "OFFICIAL EMAIL ADDRESS", "short_text", "", "Enter company official email address", "max_length", "100", ""],
        ["Field_2_1", "SEC_2", "DEPARTMENT NAME", "choice", "", "Select employee department", "none", "", "Operations, Engineering, Human Resources, Finance, IT & Systems, Sales & Marketing, Executive Management"],
        ["Field_2_2", "SEC_2", "DESIGNATION / POSITION", "short_text", "", "Enter job designation or position title", "max_length", "100", ""],
        ["Field_2_3", "SEC_2", "JOINING DATE (DD-MM-YYYY)", "date", "", "Enter company joining date in DD-MM-YYYY format", "date_ddmmyyyy", "", ""],
        ["Field_2_4", "SEC_2", "MONTHLY BASIC SALARY ($)", "decimal", "", "Enter monthly basic salary in USD", ">=", "0.0", ""],
        ["Field_3_1", "SEC_3", "RESIDENTIAL ADDRESS & REMARKS", "long_text", "", "Enter complete residential address and additional notes", "min_length", "10", ""],
        ["Field_3_2", "SEC_3", "EMERGENCY CONTACT & PHONE", "short_text", "", "Enter emergency contact person name and telephone number", "max_length", "150", ""]
    ]

    for r in schema_rows:
        ws3.append(r)

    # Format Header Row Sheet 3
    for col_idx in range(1, len(schema_headers) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.protection = prot_locked

    for row_idx in range(2, len(schema_rows) + 2):
        f_dtype = str(ws3.cell(row=row_idx, column=4).value or '').strip().lower()
        val_op = str(ws3.cell(row=row_idx, column=7).value or 'none').strip().lower()

        for col_idx in range(1, len(schema_headers) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.border = border_thin

            # Structural Columns (Field_ID [Col A], Section_ID [Col B]) -> Locked
            if col_idx in (1, 2):
                cell.font = font_bold_navy
                cell.protection = prot_locked
                if row_idx % 2 == 1:
                    cell.fill = fill_zebra
            elif col_idx == 8:  # Column H: Validation_Param_Value
                requires_param = val_op not in ('none', 'date_ddmmyyyy', '')
                if requires_param:
                    cell.font = font_regular
                    cell.fill = fill_editable
                    cell.protection = prot_unlocked
                else:
                    cell.font = font_regular
                    cell.fill = fill_zebra
                    cell.protection = prot_locked
            elif col_idx == 9:  # Column I: Dropdown_Options (Unlocked ONLY if Field_DataType is 'choice')
                if f_dtype == 'choice':
                    cell.font = font_regular
                    cell.fill = fill_editable
                    cell.protection = prot_unlocked
                else:
                    cell.font = font_regular
                    cell.fill = fill_zebra
                    cell.protection = prot_locked
            else:
                # Always Configurable Columns (Col C, D, E, F, G) -> Soft Blue Fill & Unlocked
                cell.font = font_regular
                cell.fill = fill_editable
                cell.protection = prot_unlocked

    # --- EXCEL IN-CELL DATA VALIDATIONS FOR SHEET 3 ---
    # 1. Field_DataType Dropdown Validation (PURE DATATYPES ONLY)
    dv_ftype = DataValidation(
        type="list",
        formula1='"short_text,long_text,whole_number,decimal,choice,multi_choice,radio_choice,date"',
        allow_blank=False,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Invalid Field DataType!",
        error="Invalid Field DataType! Please select short_text, long_text, whole_number, decimal, choice, or date.",
        promptTitle="Field DataType", prompt="Select field datatype."
    )
    ws3.add_data_validation(dv_ftype)
    dv_ftype.add(f"D2:D{len(schema_rows) + 1}")

    # 2. CASCADING DEPENDENT Validation_Operator Dropdown =INDIRECT(D2)
    dv_op = DataValidation(
        type="list",
        formula1='INDIRECT(D2)',
        allow_blank=False,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Invalid Validation Operator!",
        error="Selected operator is not allowed for this Field_DataType! Please choose a valid operator from the dependent dropdown.",
        promptTitle="Validation Operator", prompt="Select dependent operator allowed for the chosen Field_DataType."
    )
    ws3.add_data_validation(dv_op)
    dv_op.add(f"G2:G{len(schema_rows) + 1}")

    # 3. Validation_Param_Value Range Boundary Formula Validation (Column H: Lower < Upper for in_range / out_range)
    dv_range_param = DataValidation(
        type="custom",
        formula1='IF(OR(G2="in_range", G2="out_range"), AND(ISNUMBER(VALUE(LEFT(H2, FIND(",", H2)-1))), ISNUMBER(VALUE(MID(H2, FIND(",", H2)+1, LEN(H2)))), VALUE(LEFT(H2, FIND(",", H2)-1)) < VALUE(MID(H2, FIND(",", H2)+1, LEN(H2)))), TRUE)',
        allow_blank=True,
        showErrorMessage=True, showInputMessage=True,
        errorTitle="Invalid Range Boundaries!",
        error="For 'in_range' or 'out_range', enter 2 numbers separated by a comma where the 1st number (lower bound) is strictly smaller than the 2nd number (upper bound) (e.g. '18, 60' or '0.0, 15.0').",
        promptTitle="Range Parameter Format",
        prompt="Enter single threshold (e.g. 100) or range bounds (e.g. 18, 60 where lower < upper)."
    )
    ws3.add_data_validation(dv_range_param)
    dv_range_param.add(f"H2:H{len(schema_rows) + 1}")

    # Turn ON Sheet Protection for Sheet 3
    ws3.protection.sheet = True

    # Auto-fit columns Sheet 3
    for col in ws3.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(min(max_len + 4, 45), 14)

    # Save to Master path inside Interactive_PDF_Code/ (Catch PermissionError gracefully)
    try:
        wb.save(abs_path)
        logger.info(f"Successfully generated master 3-sheet Form Configurator Excel workbook at: '{abs_path}'.")
    except PermissionError:
        raise PermissionError(f"File Permission Error: '{abs_path}' is currently open in Microsoft Excel. Please close Microsoft Excel and try again.")

    return abs_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    path = create_3sheet_pdf_configurator_excel("Interactive_PDF_Code/fillable_pdf_configurator.xlsx")
    print(f"Master 3-Sheet Fillable PDF Form Configurator Created at: {path}")
